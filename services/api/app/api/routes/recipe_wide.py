"""Wide-table import/export for the recipe-and-brush workspace.

The generic bulk importer writes one DB table per resource. The recipe workspace
needs a single *wide* spreadsheet that fans out into SprayProgramVersion -> Brush
-> BrushParameter (EAV) -> BrushPointContribution. This module provides the
template/export/import endpoints that split/merge that wide table.
"""

from __future__ import annotations

import csv
from io import BytesIO, StringIO
from typing import Any, Literal

from fastapi import APIRouter, Depends, HTTPException, Query, Request, Response
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.session import get_db
from app.models.domain import (
    Brush,
    BrushParameter,
    BrushPointContribution,
    MeasurementPoint,
    ParameterDefinition,
    ProcessStage,
    SprayProgram,
    SprayProgramVersion,
)
from app.schemas.process import RecipeWideImportError, RecipeWideImportResult

router = APIRouter(prefix="/recipe-wide", tags=["recipe-wide"])

# ProcessStage value -> parameter_code prefix used in PARAMETER_CATALOG.
_STAGE_PREFIX: dict[str, str] = {
    ProcessStage.MIDCOAT_EXT.value: "midcoat",
    ProcessStage.BASECOAT_1.value: "basecoat_1",
    ProcessStage.BASECOAT_2.value: "basecoat_2",
    ProcessStage.CLEARCOAT_1.value: "clearcoat_1",
    ProcessStage.CLEARCOAT_2.value: "clearcoat_2",
}

# (suffix, display label) for the five brush spray parameters, in display order.
_PARAM_SUFFIXES: tuple[tuple[str, str], ...] = (
    ("spray_flow", "喷涂流量"),
    ("outer_air", "外成型空气流量"),
    ("inner_air", "内成型空气流量"),
    ("bell_speed", "旋杯转速"),
    ("voltage", "静电高压"),
)

IDENTITY_COLUMNS: tuple[str, ...] = (
    "喷涂程序",
    "版本",
    "刷子表",
    "刷子号",
    "喷涂点位",
    "测量点",
)


def _param_codes(stage: str) -> list[tuple[str, str]]:
    """Return [(parameter_code, display_label), ...] for a process stage."""
    prefix = _STAGE_PREFIX.get(stage)
    if not prefix:
        raise HTTPException(
            status_code=422,
            detail=f"工艺阶段 {stage} 不支持宽表导入（仅中涂/色漆/清漆一/二遍）",
        )
    return [(f"{prefix}_{suffix}", label) for suffix, label in _PARAM_SUFFIXES]


def _wide_columns(param_codes: list[tuple[str, str]]) -> list[str]:
    """Full ordered column list for the wide table."""
    cols: list[str] = list(IDENTITY_COLUMNS)
    for _, label in param_codes:
        cols.extend([label, f"{label}_下限", f"{label}_上限"])
    cols.append("权重")
    return cols


def _cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        # Trim trailing .0 for integer-valued floats, keep precision otherwise.
        if value.is_integer():
            return str(int(value))
        return repr(value)
    return str(value)


def _csv_bytes(columns: list[str], rows: list[dict[str, Any]]) -> bytes:
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=columns)
    writer.writeheader()
    for row in rows:
        writer.writerow({col: _cell(row.get(col)) for col in columns})
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def _xlsx_bytes(columns: list[str], rows: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "data"
    sheet.append(columns)
    for row in rows:
        sheet.append([_cell(row.get(col)) for col in columns])
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 60)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _file_response(
    columns: list[str],
    rows: list[dict[str, Any]],
    file_format: str,
    purpose: str,
) -> Response:
    if file_format == "xlsx":
        content = _xlsx_bytes(columns, rows)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        suffix = "xlsx"
    else:
        content = _csv_bytes(columns, rows)
        media_type = "text/csv; charset=utf-8"
        suffix = "csv"
    filename = f"recipe-wide-{purpose}.{suffix}"
    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _parse_rows(content: bytes, filename: str) -> list[dict[str, Any]]:
    if not content:
        raise HTTPException(status_code=422, detail="导入文件为空")
    if filename.lower().endswith(".xlsx"):
        return _parse_xlsx(content)
    return _parse_csv(content)


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(status_code=422, detail="CSV 缺少表头")
    return [dict(row) for row in reader]


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=422, detail="Excel 缺少表头") from None
    headers = [str(h).strip() if h is not None else "" for h in headers]
    result: list[dict[str, Any]] = []
    for row in rows_iter:
        if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        result.append({headers[i]: ("" if i >= len(row) or row[i] is None else row[i]) for i in range(len(headers))})
    return result


def _resolve_program(db: Session, program_id: str) -> SprayProgram:
    program = db.get(SprayProgram, program_id)
    if not program:
        raise HTTPException(status_code=404, detail="喷涂程序不存在")
    return program


def _resolve_version(db: Session, version_id: str) -> SprayProgramVersion:
    version = db.get(SprayProgramVersion, version_id)
    if not version:
        raise HTTPException(status_code=404, detail="程序版本不存在")
    return version


def _build_wide_rows(
    db: Session,
    program: SprayProgram,
    version: SprayProgramVersion,
    brush_table_no: str | None,
    param_codes: list[tuple[str, str]],
) -> list[dict[str, Any]]:
    """Build wide rows from existing brush/parameter/contribution data."""
    brush_query = select(Brush).where(Brush.program_version_id == version.id)
    if brush_table_no:
        brush_query = brush_query.where(Brush.brush_table_no == brush_table_no)
    brush_query = brush_query.order_by(Brush.brush_no)
    brushes = list(db.scalars(brush_query))
    if not brushes:
        return []

    brush_ids = [b.id for b in brushes]
    params_by_brush: dict[str, dict[str, BrushParameter]] = {}
    for param in db.scalars(select(BrushParameter).where(BrushParameter.brush_id.in_(brush_ids))):
        params_by_brush.setdefault(param.brush_id, {})[param.parameter_code] = param

    contribs_by_brush: dict[str, list[BrushPointContribution]] = {}
    for contrib in db.scalars(
        select(BrushPointContribution).where(BrushPointContribution.brush_id.in_(brush_ids))
    ):
        contribs_by_brush.setdefault(contrib.brush_id, []).append(contrib)

    point_ids = {c.measurement_point_id for contribs in contribs_by_brush.values() for c in contribs}
    point_code_by_id: dict[str, str] = {}
    if point_ids:
        for point in db.scalars(select(MeasurementPoint).where(MeasurementPoint.id.in_(point_ids))):
            point_code_by_id[point.id] = point.code

    rows: list[dict[str, Any]] = []
    for brush in brushes:
        brush_params = params_by_brush.get(brush.id, {})
        contribs = contribs_by_brush.get(brush.id, [])
        point_rows = contribs or [None]
        for contrib in point_rows:
            row: dict[str, Any] = {
                "喷涂程序": program.program_code,
                "版本": version.version,
                "刷子表": brush.brush_table_no,
                "刷子号": brush.brush_no,
                "喷涂点位": brush.spray_position or "",
                "测量点": point_code_by_id.get(contrib.measurement_point_id, "") if contrib else "",
            }
            for code, label in param_codes:
                bp = brush_params.get(code)
                row[label] = bp.configured_value if bp else ""
                row[f"{label}_下限"] = bp.soft_min if bp else ""
                row[f"{label}_上限"] = bp.soft_max if bp else ""
            row["权重"] = contrib.contribution_weight if contrib else ""
            rows.append(row)
    return rows


@router.get("/template")
def recipe_wide_template(
    spray_program_id: str = Query(...),
    program_version_id: str = Query(...),
    brush_table_no: str | None = Query(default=None),
    file_format: Literal["csv", "xlsx"] = Query(default="xlsx", alias="format"),
    db: Session = Depends(get_db),
) -> Response:
    program = _resolve_program(db, spray_program_id)
    version = _resolve_version(db, program_version_id)
    param_codes = _param_codes(program.process_stage)
    columns = _wide_columns(param_codes)
    rows = _build_wide_rows(db, program, version, brush_table_no, param_codes)
    if not rows:
        # Emit a single prefilled blank row so the user sees the identity defaults.
        blank: dict[str, Any] = {col: "" for col in columns}
        blank["喷涂程序"] = program.program_code
        blank["版本"] = version.version
        if brush_table_no:
            blank["刷子表"] = brush_table_no
        rows = [blank]
    return _file_response(columns, rows, file_format, purpose="template")


@router.get("/export")
def recipe_wide_export(
    spray_program_id: str = Query(...),
    program_version_id: str = Query(...),
    brush_table_no: str | None = Query(default=None),
    file_format: Literal["csv", "xlsx"] = Query(default="xlsx", alias="format"),
    db: Session = Depends(get_db),
) -> Response:
    program = _resolve_program(db, spray_program_id)
    version = _resolve_version(db, program_version_id)
    param_codes = _param_codes(program.process_stage)
    columns = _wide_columns(param_codes)
    rows = _build_wide_rows(db, program, version, brush_table_no, param_codes)
    return _file_response(columns, rows, file_format, purpose="export")


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        raise ValueError(f"无法解析数值: {value!r}")


def _param_def_by_code(db: Session, code: str) -> ParameterDefinition | None:
    return db.scalar(select(ParameterDefinition).where(ParameterDefinition.code == code))


def _upsert_brush(
    db: Session, version_id: str, brush_no: str, brush_table_no: str, spray_position: str | None
) -> tuple[Brush, bool]:
    brush = db.scalar(
        select(Brush).where(
            Brush.program_version_id == version_id,
            Brush.brush_no == brush_no,
        )
    )
    if brush:
        if brush.brush_table_no != brush_table_no:
            brush.brush_table_no = brush_table_no
        if spray_position is not None and spray_position != "":
            brush.spray_position = spray_position
        db.flush()
        return brush, False
    brush = Brush(
        program_version_id=version_id,
        brush_no=brush_no,
        brush_table_no=brush_table_no,
        spray_position=spray_position or None,
    )
    db.add(brush)
    db.flush()
    return brush, True


def _upsert_brush_parameter(
    db: Session,
    brush_id: str,
    code: str,
    label: str,
    configured_value: float | None,
    soft_min: float | None,
    soft_max: float | None,
) -> None:
    if configured_value is None:
        return
    param = db.scalar(
        select(BrushParameter).where(
            BrushParameter.brush_id == brush_id,
            BrushParameter.parameter_code == code,
        )
    )
    if param:
        param.configured_value = configured_value
        param.soft_min = soft_min
        param.soft_max = soft_max
        db.flush()
        return
    definition = _param_def_by_code(db, code)
    param = BrushParameter(
        brush_id=brush_id,
        parameter_definition_id=definition.id if definition else None,
        parameter_code=code,
        parameter_name=label,
        configured_value=configured_value,
        unit="",
        soft_min=soft_min,
        soft_max=soft_max,
        is_recommendable=False,
    )
    db.add(param)
    db.flush()


def _resolve_point_by_code(db: Session, code: str) -> MeasurementPoint | None:
    return db.scalar(select(MeasurementPoint).where(MeasurementPoint.code == code.strip()))


@router.post("/import", response_model=RecipeWideImportResult)
async def recipe_wide_import(
    request: Request,
    spray_program_id: str = Query(...),
    program_version_id: str = Query(...),
    brush_table_no: str | None = Query(default=None),
    filename: str = Query(default="recipe-wide-import.csv"),
    db: Session = Depends(get_db),
) -> RecipeWideImportResult:
    program = _resolve_program(db, spray_program_id)
    version = _resolve_version(db, program_version_id)
    if version.spray_program_id != program.id:
        raise HTTPException(status_code=422, detail="版本不属于所选喷涂程序")
    param_codes = _param_codes(program.process_stage)
    label_to_code = {label: code for code, label in param_codes}

    rows = _parse_rows(await request.body(), filename)
    errors: list[RecipeWideImportError] = []
    created = 0
    updated = 0
    skipped = 0

    for index, row in enumerate(rows, start=1):
        try:
            brush_no = str(row.get("刷子号", "")).strip()
            if not brush_no:
                skipped += 1
                continue
            row_brush_table = str(row.get("刷子表", "")).strip() or brush_table_no or ""
            if not row_brush_table:
                raise ValueError("缺少刷子表（列「刷子表」或查询参数 brush_table_no）")
            spray_position = str(row.get("喷涂点位", "")).strip() or None

            brush, is_new_brush = _upsert_brush(db, version.id, brush_no, row_brush_table, spray_position)

            for label, code in label_to_code.items():
                configured_value = _to_float(row.get(label))
                soft_min = _to_float(row.get(f"{label}_下限"))
                soft_max = _to_float(row.get(f"{label}_上限"))
                if configured_value is None and soft_min is None and soft_max is None:
                    continue
                _upsert_brush_parameter(db, brush.id, code, label, configured_value, soft_min, soft_max)

            point_code = str(row.get("测量点", "")).strip()
            weight_raw = row.get("权重")
            if point_code and weight_raw not in (None, ""):
                point = _resolve_point_by_code(db, point_code)
                if not point:
                    raise ValueError(f"测量点 {point_code} 不存在")
                weight = _to_float(weight_raw)
                if weight is None:
                    raise ValueError("权重不是合法数值")
                contribution = db.scalar(
                    select(BrushPointContribution).where(
                        BrushPointContribution.brush_id == brush.id,
                        BrushPointContribution.measurement_point_id == point.id,
                    )
                )
                if contribution:
                    contribution.contribution_weight = weight
                    updated += 1
                else:
                    db.add(
                        BrushPointContribution(
                            brush_id=brush.id,
                            measurement_point_id=point.id,
                            overlap_ratio=0.0,
                            contribution_weight=weight,
                        )
                    )
                    created += 1
                db.flush()
            elif is_new_brush:
                created += 1
            else:
                updated += 1
            db.commit()
        except ValueError as exc:
            db.rollback()
            errors.append(RecipeWideImportError(row=index, message=str(exc)))
        except HTTPException as exc:
            db.rollback()
            errors.append(RecipeWideImportError(row=index, message=exc.detail))
        except Exception as exc:  # noqa: BLE001 - surface any row failure to the user
            db.rollback()
            errors.append(RecipeWideImportError(row=index, message=f"未知错误: {exc}"))

    return RecipeWideImportResult(
        total_rows=len(rows),
        created=created,
        updated=updated,
        skipped=skipped,
        failed=len(errors),
        errors=errors,
    )



