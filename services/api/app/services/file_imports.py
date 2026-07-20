from __future__ import annotations

import base64
import csv
import hashlib
import json
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO, StringIO
from typing import Any, Callable

from fastapi import HTTPException
from openpyxl import load_workbook

from app.models.domain import FileImportProfile

MAX_PERSISTED_IMPORT_ROWS = 20_000
TARGET_RESOURCE_ALIASES = {
    "quality_measurement": "quality.measurements",
    "supplier-submissions": "engineering.supplier-submissions",
    "trajectory-geometries": "engineering.trajectory-geometries",
}


@dataclass(frozen=True)
class ImportPreview:
    row_count: int
    valid_row_count: int
    failed_row_count: int
    preview_payload: dict[str, Any]
    error_report: dict[str, Any] | None
    source_checksum: str


def decode_base64_file(content_base64: str, *, max_bytes: int | None = None) -> bytes:
    if "," in content_base64 and content_base64.split(",", 1)[0].startswith("data:"):
        content_base64 = content_base64.split(",", 1)[1]
    try:
        content = base64.b64decode(content_base64, validate=True)
    except Exception as exc:  # noqa: BLE001 - normalize parser errors for API clients.
        raise HTTPException(status_code=422, detail="文件内容不是有效 base64") from exc
    if not content:
        raise HTTPException(status_code=422, detail="导入文件为空")
    if max_bytes is not None and len(content) > max_bytes:
        raise HTTPException(status_code=413, detail="导入文件超过系统允许的大小")
    return content


def resolve_target_resource(target_resource: str) -> str:
    return TARGET_RESOURCE_ALIASES.get(target_resource, target_resource)


def build_import_preview(
    profile: FileImportProfile,
    content: bytes,
    *,
    source_filename: str,
    source_checksum: str | None = None,
    preview_limit: int = 20,
) -> ImportPreview:
    rows = _parse_rows(content, source_filename, profile.parser_type)
    rules = profile.validation_rules or {}
    max_rows = rules.get("max_rows")
    errors: list[dict[str, Any]] = []
    if isinstance(max_rows, int) and max_rows >= 0 and len(rows) > max_rows:
        errors.append(
            {
                "row": 0,
                "field": "file",
                "message": f"文件行数 {len(rows)} 超过 profile 限制 {max_rows}",
            }
        )
    if len(rows) > MAX_PERSISTED_IMPORT_ROWS:
        errors.append(
            {
                "row": 0,
                "field": "file",
                "message": f"文件行数超过单任务上限 {MAX_PERSISTED_IMPORT_ROWS}",
            }
        )

    mapped_rows: list[dict[str, Any]] = []
    source_headers = sorted({key for row in rows for key in row})
    for index, row in enumerate(rows, start=2):
        mapped = _map_row(row, profile.field_mapping or {})
        mapped_rows.append(mapped)
        errors.extend(_validate_row(mapped, index, profile.required_fields or [], rules))
    errors.extend(_validate_target_rows(profile.target_resource, mapped_rows))

    has_file_error = any(error.get("row") == 0 for error in errors)
    failed_row_numbers = {
        int(error["row"])
        for error in errors
        if isinstance(error.get("row"), int) and int(error["row"]) > 0
    }
    failed_row_count = len(rows) if has_file_error else len(failed_row_numbers)
    preview_payload = {
        "profile_id": profile.id,
        "profile_code": profile.code,
        "profile_version": profile.version,
        "domain_type": profile.domain_type,
        "parser_type": profile.parser_type,
        "target_resource": profile.target_resource,
        "resolved_target_resource": resolve_target_resource(profile.target_resource),
        "source_filename": source_filename,
        "source_headers": source_headers,
        "target_fields": sorted({key for row in mapped_rows for key in row}),
        "field_mapping": profile.field_mapping or {},
        "required_fields": profile.required_fields or [],
        "validation_rules": rules,
        "row_count": len(mapped_rows),
        "preview_rows": mapped_rows[: max(preview_limit, 0)],
        "validated_rows": mapped_rows if len(mapped_rows) <= MAX_PERSISTED_IMPORT_ROWS else [],
        "truncated_preview": len(mapped_rows) > max(preview_limit, 0),
        "validation_status": "PASSED" if not errors else "FAILED",
    }
    error_report = {
        "error_count": len(errors),
        "errors": errors[:200],
        "truncated_errors": len(errors) > 200,
    } if errors else None
    return ImportPreview(
        row_count=len(rows),
        valid_row_count=max(0, len(rows) - failed_row_count),
        failed_row_count=failed_row_count,
        preview_payload=preview_payload,
        error_report=error_report,
        source_checksum=source_checksum or hashlib.sha256(content).hexdigest(),
    )


def execute_validated_import(
    profile: FileImportProfile,
    preview_payload: dict[str, Any] | None,
    *,
    source_filename: str,
    mode: str,
    heartbeat: Callable[[], None] | None = None,
    db,
) -> dict[str, Any]:
    if not preview_payload or preview_payload.get("validation_status") != "PASSED":
        raise HTTPException(status_code=409, detail="导入任务未通过预览校验，不能写入业务数据")
    rows = preview_payload.get("validated_rows")
    if not isinstance(rows, list):
        raise HTTPException(status_code=409, detail="导入任务缺少可重放的规范化数据")
    if len(rows) != int(preview_payload.get("row_count", len(rows))):
        raise HTTPException(status_code=409, detail="导入任务规范化数据不完整，请重新上传并校验")
    preview_target = preview_payload.get("resolved_target_resource")
    current_target = resolve_target_resource(profile.target_resource)
    if not isinstance(preview_target, str) or not preview_target:
        raise HTTPException(status_code=409, detail="导入预览缺少冻结的目标资源，请重新上传并校验")
    if preview_target != current_target:
        raise HTTPException(status_code=409, detail="导入配置的目标资源已变化，请重新上传并校验")

    from app.services.bulk_io import import_resource

    if heartbeat:
        heartbeat()
    result = import_resource(
        preview_target,
        _rows_to_csv(rows),
        filename=f"{source_filename}.normalized.csv",
        mode=mode,
        progress_callback=heartbeat,
        db=db,
    )
    if heartbeat:
        heartbeat()
    return result


def _validate_target_rows(
    target_resource: str,
    rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    from app.services.bulk_io import get_resource

    try:
        resource = get_resource(resolve_target_resource(target_resource))
    except HTTPException as exc:
        return [{"row": 0, "field": "target_resource", "message": str(exc.detail)}]
    if not resource.importable:
        return [{"row": 0, "field": "target_resource", "message": "目标资源不允许批量导入"}]

    allowed = {field.name for field in resource.fields if field.name != "id"}
    required = {field.name for field in resource.fields if field.required and field.name != "id"}
    errors: list[dict[str, Any]] = []
    unknown = sorted({field for row in rows for field in row if field not in allowed})
    if unknown:
        errors.append(
            {
                "row": 0,
                "field": "field_mapping",
                "message": f"映射后的字段不属于目标资源：{', '.join(unknown)}",
            }
        )
    for row_index, row in enumerate(rows, start=2):
        missing = sorted(field for field in required if _blank(row.get(field)))
        if missing:
            errors.append(
                {
                    "row": row_index,
                    "field": ",".join(missing),
                    "message": "目标资源必填字段缺失",
                }
            )
    return errors


def _rows_to_csv(rows: list[dict[str, Any]]) -> bytes:
    headers = sorted({str(key) for row in rows for key in row})
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    for row in rows:
        writer.writerow(
            {
                header: json.dumps(row.get(header), ensure_ascii=False, default=str)
                if isinstance(row.get(header), (dict, list))
                else row.get(header)
                for header in headers
            }
        )
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def _parse_rows(content: bytes, filename: str, parser_type: str) -> list[dict[str, Any]]:
    lowered = filename.lower()
    if parser_type == "XLSX" or lowered.endswith(".xlsx"):
        return _parse_xlsx(content)
    if parser_type in {"CSV", "DXQ_EXPORT"} or lowered.endswith(".csv"):
        return _parse_csv(content)
    if parser_type in {"JSON", "XML"}:
        raise HTTPException(status_code=422, detail=f"当前导入预览暂不支持 {parser_type} 解析")
    raise HTTPException(status_code=422, detail=f"不支持的导入文件类型：{filename}")


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("gb18030")
    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(status_code=422, detail="CSV 缺少表头")
    return [
        {str(key).strip(): value for key, value in row.items() if key is not None and str(key).strip()}
        for row in reader
    ]


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook["data"] if "data" in workbook.sheetnames else workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=422, detail="Excel 缺少表头")
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    if not any(headers):
        raise HTTPException(status_code=422, detail="Excel 表头为空")
    return [
        {
            headers[index]: values[index] if index < len(values) else None
            for index in range(len(headers))
            if headers[index]
        }
        for values in rows[1:]
    ]


def _map_row(row: dict[str, Any], mapping: dict[str, Any]) -> dict[str, Any]:
    mapped: dict[str, Any] = {}
    for source, value in row.items():
        target = mapping.get(source, source)
        if isinstance(target, str) and target:
            mapped[target] = value
    return mapped


def _validate_row(
    row: dict[str, Any],
    row_index: int,
    required_fields: list[str],
    rules: dict[str, Any],
) -> list[dict[str, Any]]:
    errors: list[dict[str, Any]] = []
    for field in required_fields:
        if _blank(row.get(field)):
            errors.append({"row": row_index, "field": field, "message": "缺少必填字段"})

    required_any = rules.get("required_any") if isinstance(rules.get("required_any"), list) else []
    for group in required_any:
        if isinstance(group, list) and not any(not _blank(row.get(str(field))) for field in group):
            errors.append(
                {
                    "row": row_index,
                    "field": ",".join(str(field) for field in group),
                    "message": "至少需要填写其中一个字段",
                }
            )

    for field in _rule_fields(rules, "numeric_fields"):
        if not _blank(row.get(field)):
            try:
                float(row[field])
            except (TypeError, ValueError):
                errors.append({"row": row_index, "field": field, "message": "字段必须是数值"})

    for field in _rule_fields(rules, "integer_fields"):
        if not _blank(row.get(field)):
            try:
                int(str(row[field]))
            except (TypeError, ValueError):
                errors.append({"row": row_index, "field": field, "message": "字段必须是整数"})

    for field in _rule_fields(rules, "datetime_fields"):
        if not _blank(row.get(field)):
            try:
                datetime.fromisoformat(str(row[field]).replace("Z", "+00:00"))
            except ValueError:
                errors.append({"row": row_index, "field": field, "message": "字段必须是 ISO 日期时间"})

    allowed_values = rules.get("allowed_values")
    if isinstance(allowed_values, dict):
        for field, allowed in allowed_values.items():
            if _blank(row.get(field)) or not isinstance(allowed, list):
                continue
            if str(row[field]) not in {str(item) for item in allowed}:
                errors.append(
                    {
                        "row": row_index,
                        "field": field,
                        "message": f"字段值不在允许范围：{', '.join(str(item) for item in allowed)}",
                    }
                )
    return errors


def _rule_fields(rules: dict[str, Any], key: str) -> list[str]:
    values = rules.get(key)
    if not isinstance(values, list):
        return []
    return [str(value) for value in values]


def _blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())
