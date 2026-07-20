from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO, StringIO
from types import UnionType
from typing import Any, Callable, Literal, Union, get_args, get_origin

from fastapi import HTTPException
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field, ValidationError
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.api.routes.body_map import upsert_body_map_3d_layout, upsert_body_map_layout
from app.api.routes.factories import create_factory, update_factory
from app.api.routes.engineering import (
    create_contribution_validation,
    reject_direct_file_import_job_create,
    create_file_import_profile,
    create_issue_task,
    create_issue_task_comment,
    create_issue_task_evidence,
    create_knowledge_entry,
    create_measurement_msa_study,
    create_measurement_probe,
    create_model_explanation,
    create_process_route,
    create_process_route_applicability,
    create_process_route_step,
    create_supplier_issue,
    create_supplier_submission,
    create_trajectory_geometry,
    update_contribution_validation,
    update_file_import_job,
    update_file_import_profile,
    update_issue_task,
    update_knowledge_entry,
    update_measurement_msa_study,
    update_measurement_probe,
    update_process_route,
    update_process_route_applicability,
    update_process_route_step,
    update_supplier_issue,
    update_supplier_submission,
    update_trajectory_geometry,
)
from app.api.routes.integration import create_endpoint, create_event, update_endpoint
from app.api.routes.master_data import (
    bind_factory_vehicle_model,
    bind_measurement_group_point,
    bind_vehicle_model_color,
    create_color,
    create_measurement_group,
    create_measurement_point,
    create_part,
    create_vehicle_model,
    update_color,
    update_measurement_group,
    update_measurement_point,
    update_part,
    update_vehicle_model,
)
from app.api.routes.material_governance import (
    create_applicability as create_material_applicability,
    create_definition as create_material_definition,
    create_method as create_material_method,
    create_result as create_material_result,
    create_specification as create_material_specification,
    update_applicability as update_material_applicability,
    update_definition as update_material_definition,
    update_method as update_material_method,
    update_result as update_material_result,
    update_specification as update_material_specification,
)
from app.api.routes.measurement_governance import (
    create_calibration,
    create_import_profile,
    create_instrument,
    create_method as create_measurement_method,
    create_reference,
    update_calibration,
    update_import_profile,
    update_instrument,
    update_method as update_measurement_method,
    update_reference,
)
from app.api.routes.process import (
    create_actual_parameter,
    create_brush,
    create_brush_parameter,
    create_material_batch,
    create_parameter_constraint_source,
    create_parameter_definition,
    create_production_run,
    create_production_stage_run,
    create_program_version,
    create_spray_program,
    update_actual_parameter,
    update_brush,
    update_brush_parameter,
    update_material_batch,
    update_parameter_constraint_source,
    update_production_run,
    update_production_stage_run,
    update_program_version,
    update_spray_program,
    upsert_brush_point_contribution,
)
from app.api.routes.quality import (
    create_quality_measurement,
    create_quality_standard,
    update_quality_measurement,
    update_quality_standard,
)
from app.api.routes.robot_governance import (
    create_atomizer,
    create_contribution_entry,
    create_contribution_version,
    create_controller,
    create_device_configuration,
    create_device_execution,
    create_path_segment,
    create_robot,
    create_trajectory_program,
    update_atomizer,
    update_contribution_entry,
    update_contribution_version,
    update_controller,
    update_device_configuration,
    update_device_execution,
    update_path_segment,
    update_robot,
    update_trajectory_program,
)
from app.core.referential_integrity import check_fk
from app.domain.quality_metric_catalog import QUALITY_METRIC_CATALOG
from app.models.domain import (
    ActualParameter,
    AppUser,
    Brush,
    BrushParameter,
    BrushPointContribution,
    Color,
    ContributionValidationStudy,
    DurrApplicationController,
    DurrRobot,
    DurrRotaryAtomizer,
    EngineeringKnowledgeEntry,
    Factory,
    FactoryVehicleModel,
    FileImportJob,
    FileImportProfile,
    IntegrationEndpoint,
    IntegrationEvent,
    MaterialBatch,
    MaterialBatchTestResult,
    MaterialCharacteristicApplicability,
    MaterialCharacteristicDefinition,
    MaterialSpecification,
    MaterialTestMethod,
    MeasurementCalibrationRecord,
    MeasurementGroup,
    MeasurementGroupPoint,
    MeasurementImportProfile,
    MeasurementInstrument,
    MeasurementMethod,
    MeasurementMsaStudy,
    MeasurementPoint,
    MeasurementPoint3DLayout,
    MeasurementPointLayout,
    MeasurementProbe,
    MeasurementReferenceStandard,
    ModelExplanation,
    MeasurementRepeatReading,
    ParameterConstraintSource,
    ParameterDefinition,
    Part,
    PointContributionEntry,
    PointContributionVersion,
    ProcessRoute,
    ProcessRouteApplicability,
    ProcessRouteStep,
    ProductionDeviceExecution,
    ProductionRun,
    ProductionStageRun,
    ProgramColor,
    ProgramDeviceConfiguration,
    ProgramVehicleModel,
    QualityMeasurement,
    QualityIssueComment,
    QualityIssueEvidence,
    QualityIssueTask,
    QualityMetricValue,
    QualityStandard,
    SprayProgram,
    SprayProgramVersion,
    SupplierMaterialIssue,
    SupplierMaterialSubmission,
    TrajectoryPathSegment,
    TrajectorySegmentGeometry,
    TrajectoryProgram,
    VehicleModel,
    VehicleModelColor,
)
from app.schemas.common import FactoryCreate, FactoryUpdate
from app.schemas.engineering import (
    ContributionValidationStudyCreate,
    ContributionValidationStudyUpdate,
    EngineeringKnowledgeEntryCreate,
    EngineeringKnowledgeEntryUpdate,
    FileImportJobCreate,
    FileImportJobUpdate,
    FileImportProfileCreate,
    FileImportProfileUpdate,
    MeasurementMsaStudyCreate,
    MeasurementMsaStudyUpdate,
    MeasurementProbeCreate,
    MeasurementProbeUpdate,
    ModelExplanationCreate,
    ProcessRouteApplicabilityCreate,
    ProcessRouteApplicabilityUpdate,
    ProcessRouteCreate,
    ProcessRouteStepCreate,
    ProcessRouteStepUpdate,
    ProcessRouteUpdate,
    QualityIssueTaskCreate,
    QualityIssueTaskUpdate,
    QualityIssueCommentCreate,
    QualityIssueEvidenceCreate,
    SupplierMaterialIssueCreate,
    SupplierMaterialIssueUpdate,
    SupplierMaterialSubmissionCreate,
    SupplierMaterialSubmissionUpdate,
    TrajectorySegmentGeometryCreate,
    TrajectorySegmentGeometryUpdate,
)
from app.schemas.integration import (
    IntegrationEndpointCreate,
    IntegrationEndpointUpdate,
    IntegrationEventCreate,
)
from app.schemas.master_data import (
    ColorCreate,
    ColorUpdate,
    FactoryVehicleModelCreate,
    MeasurementGroupCreate,
    MeasurementGroupPointBind,
    MeasurementGroupUpdate,
    MeasurementPointCreate,
    MeasurementPointUpdate,
    PartCreate,
    PartUpdate,
    VehicleModelColorCreate,
    VehicleModelCreate,
    VehicleModelUpdate,
)
from app.schemas.material import (
    MaterialBatchTestResultCreate,
    MaterialBatchTestResultUpdate,
    MaterialCharacteristicApplicabilityCreate,
    MaterialCharacteristicApplicabilityUpdate,
    MaterialCharacteristicDefinitionCreate,
    MaterialCharacteristicDefinitionUpdate,
    MaterialSpecificationCreate,
    MaterialSpecificationUpdate,
    MaterialTestMethodCreate,
    MaterialTestMethodUpdate,
)
from app.schemas.process import (
    ActualParameterCreate,
    ActualParameterUpdate,
    BrushCreate,
    BrushParameterCreate,
    BrushParameterUpdate,
    BrushPointContributionUpsert,
    BrushUpdate,
    DurrAtomizerCreate,
    DurrAtomizerUpdate,
    DurrControllerCreate,
    DurrControllerUpdate,
    DurrRobotCreate,
    DurrRobotUpdate,
    MaterialBatchCreate,
    MaterialBatchUpdate,
    ParameterConstraintSourceCreate,
    ParameterConstraintSourceUpdate,
    ParameterDefinitionCreate,
    PointContributionEntryCreate,
    PointContributionEntryUpdate,
    PointContributionVersionCreate,
    PointContributionVersionUpdate,
    ProductionDeviceExecutionCreate,
    ProductionDeviceExecutionUpdate,
    ProductionRunCreate,
    ProductionRunUpdate,
    ProductionStageRunCreate,
    ProductionStageRunUpdate,
    ProgramDeviceConfigurationCreate,
    ProgramDeviceConfigurationUpdate,
    SprayProgramCreate,
    SprayProgramUpdate,
    SprayProgramVersionCreate,
    SprayProgramVersionUpdate,
    TrajectoryPathSegmentCreate,
    TrajectoryPathSegmentUpdate,
    TrajectoryProgramCreate,
    TrajectoryProgramUpdate,
)
from app.schemas.quality import (
    BodyMap3DLayoutUpsert,
    BodyMapLayoutUpsert,
    MeasurementCalibrationCreate,
    MeasurementCalibrationUpdate,
    MeasurementImportProfileCreate,
    MeasurementImportProfileUpdate,
    MeasurementInstrumentCreate,
    MeasurementInstrumentUpdate,
    MeasurementMethodCreate,
    MeasurementMethodUpdate,
    MeasurementReferenceStandardCreate,
    MeasurementReferenceStandardUpdate,
    QualityMeasurementCreate,
    QualityMeasurementUpdate,
    QualityStandardCreate,
    QualityStandardUpdate,
)

FileFormat = Literal["csv", "xlsx"]
ImportMode = Literal["create", "upsert"]


class ProgramVehicleModelBulkCreate(BaseModel):
    program_version_id: str = Field(min_length=1, max_length=36)
    vehicle_model_id: str = Field(min_length=1, max_length=36)


class ProgramColorBulkCreate(BaseModel):
    program_version_id: str = Field(min_length=1, max_length=36)
    color_id: str = Field(min_length=1, max_length=36)


@dataclass(frozen=True)
class BulkField:
    name: str
    label: str
    kind: str
    required: bool = False
    example: str = ""
    description: str = ""


@dataclass(frozen=True)
class BulkResource:
    key: str
    label: str
    group: str
    model: type
    create_schema: type[BaseModel]
    update_schema: type[BaseModel] | None
    create_func: Callable[..., Any]
    update_func: Callable[..., Any] | None = None
    create_arg_fields: tuple[str, ...] = ()
    update_arg_fields: tuple[str, ...] = ()
    update_uses_resource_id: bool = True
    unique_fields: tuple[str, ...] = ()
    order_by: tuple[str, ...] = ("created_at",)
    extra_fields: tuple[BulkField, ...] = ()
    export_row: Callable[[Session, Any], dict[str, Any]] | None = None
    match_existing: Callable[[Session, dict[str, Any]], str | None] | None = None
    importable: bool = True
    fields: tuple[BulkField, ...] = field(init=False)

    def __post_init__(self) -> None:
        object.__setattr__(
            self,
            "fields",
            _build_fields(self.create_schema, self.update_schema, self.extra_fields),
        )


@dataclass(frozen=True)
class BusinessReference:
    model: type
    label: str
    lookup_fields: tuple[str, ...]
    display_fields: tuple[str, ...] = ()


_BUSINESS_REFERENCES: dict[str, BusinessReference] = {
    "factory_id": BusinessReference(Factory, "工厂", ("code", "name"), ("code", "name")),
    "vehicle_model_id": BusinessReference(VehicleModel, "车型", ("code", "name"), ("code", "name")),
    "color_id": BusinessReference(Color, "颜色", ("code", "name"), ("code", "name")),
    "part_id": BusinessReference(Part, "零件", ("code", "name"), ("code", "name")),
    "measurement_group_id": BusinessReference(MeasurementGroup, "测量编组", ("code", "name"), ("code", "name")),
    "measurement_point_id": BusinessReference(MeasurementPoint, "测量点", ("code", "name"), ("code", "name")),
    "spray_program_id": BusinessReference(SprayProgram, "喷涂程序", ("program_code", "name"), ("program_code", "name")),
    "program_version_id": BusinessReference(SprayProgramVersion, "喷涂程序版本", ("version",), ("version",)),
    "brush_id": BusinessReference(Brush, "刷子号", ("brush_no", "brush_table_no"), ("brush_no", "brush_table_no")),
    "parameter_definition_id": BusinessReference(ParameterDefinition, "工艺参数", ("code", "name"), ("code", "name")),
    "material_batch_id": BusinessReference(MaterialBatch, "材料批次", ("batch_no", "material_code", "material_name"), ("batch_no", "material_name")),
    "production_run_id": BusinessReference(ProductionRun, "生产记录", ("run_no", "body_no"), ("run_no", "body_no")),
    "production_stage_run_id": BusinessReference(ProductionStageRun, "生产工序记录", ("process_stage",), ("process_stage",)),
    "instrument_id": BusinessReference(MeasurementInstrument, "测量仪器", ("code", "serial_no", "name"), ("code", "name")),
    "measurement_probe_id": BusinessReference(MeasurementProbe, "测量探头", ("code", "serial_no", "name"), ("code", "name")),
    "measurement_method_id": BusinessReference(MeasurementMethod, "测量方法", ("code", "name", "version"), ("code", "version")),
    "calibration_record_id": BusinessReference(MeasurementCalibrationRecord, "校准记录", ("calibration_no",), ("calibration_no",)),
    "reference_standard_id": BusinessReference(MeasurementReferenceStandard, "参考件", ("code", "name", "serial_no"), ("code", "name")),
    "import_profile_id": BusinessReference(MeasurementImportProfile, "导入规则", ("code", "name", "version"), ("code", "version")),
    "process_route_id": BusinessReference(ProcessRoute, "工艺路线", ("route_code", "name", "version"), ("route_code", "version")),
    "task_id": BusinessReference(QualityIssueTask, "问题任务", ("task_no", "title"), ("task_no", "title")),
    "quality_measurement_id": BusinessReference(QualityMeasurement, "质量检测记录", ("data_no",), ("data_no",)),
    "characteristic_definition_id": BusinessReference(MaterialCharacteristicDefinition, "材料特性", ("code", "name"), ("code", "name")),
    "robot_id": BusinessReference(DurrRobot, "机器人", ("code", "name"), ("code", "name")),
    "controller_id": BusinessReference(DurrApplicationController, "应用控制器", ("code", "name"), ("code", "name")),
    "atomizer_id": BusinessReference(DurrRotaryAtomizer, "静电旋杯", ("code", "name"), ("code", "name")),
    "trajectory_program_id": BusinessReference(TrajectoryProgram, "轨迹程序", ("trajectory_code", "name", "version"), ("trajectory_code", "version")),
    "path_segment_id": BusinessReference(TrajectoryPathSegment, "轨迹段", ("segment_no", "name"), ("segment_no", "name")),
    "contribution_version_id": BusinessReference(PointContributionVersion, "贡献关系版本", ("version",), ("version",)),
    "endpoint_id": BusinessReference(IntegrationEndpoint, "对接系统", ("code", "name"), ("code", "name")),
    "owner_user_id": BusinessReference(AppUser, "负责人", ("username", "display_name", "email"), ("username", "display_name")),
    "profile_id": BusinessReference(FileImportProfile, "文件导入规则", ("code", "name", "version"), ("code", "version")),
    "source_import_job_id": BusinessReference(FileImportJob, "来源导入任务", ("import_no",), ("import_no",)),
    "replay_of_job_id": BusinessReference(FileImportJob, "原导入任务", ("import_no",), ("import_no",)),
}


def _business_reference_for(resource: BulkResource, field_name: str) -> BusinessReference | None:
    if field_name == "method_id":
        if resource.key.startswith("material-governance."):
            return BusinessReference(MaterialTestMethod, "材料检测方法", ("code", "name", "version"), ("code", "version"))
        return BusinessReference(MeasurementMethod, "测量方法", ("code", "name", "version"), ("code", "version"))
    return _BUSINESS_REFERENCES.get(field_name)


def list_bulk_resources() -> list[dict[str, str | bool]]:
    return [
        {
            "key": resource.key,
            "label": resource.label,
            "group": resource.group,
            "importable": resource.importable,
        }
        for resource in RESOURCES.values()
    ]


def describe_bulk_columns(
    resource_key: str,
    *,
    quality_type: str | None = None,
) -> list[dict[str, str | bool]]:
    resource = get_resource(resource_key)
    if resource.key == "quality.measurements":
        fields = _quality_measurement_template_fields(quality_type)
    elif resource.key == "process.brush-contributions":
        fields = _brush_contribution_template_fields()
    else:
        fields = _upload_template_fields(resource.fields)
    return [
        {
            "key": field.name,
            "label": _display_field_label(field),
            "kind": field.kind,
            "required": field.required,
            "description": field.description,
        }
        for field in fields
    ]


def get_resource(resource_key: str) -> BulkResource:
    resource = RESOURCES.get(resource_key)
    if not resource:
        raise HTTPException(status_code=404, detail=f"不支持的批量资源：{resource_key}")
    return resource


def render_template(
    resource_key: str,
    file_format: FileFormat,
    *,
    db: Session | None = None,
    quality_type: str | None = None,
    factory_code: str | None = None,
    color_code: str | None = None,
    vehicle_model_code: str | None = None,
    shift: str | None = None,
    brush_id: str | None = None,
    default_values: dict[str, Any] | None = None,
) -> Response:
    resource = get_resource(resource_key)
    if resource.key == "quality.measurements":
        fields = _quality_measurement_template_fields(quality_type)
        rows = _quality_measurement_template_rows(
            db,
            quality_type=quality_type,
            factory_code=factory_code,
            color_code=color_code,
            vehicle_model_code=vehicle_model_code,
            shift=shift,
        )
        return _file_response(
            resource,
            file_format,
            rows,
            purpose="template",
            include_metadata=True,
            fields=fields,
        )
    if resource.key == "process.brush-contributions":
        fields = _brush_contribution_template_fields()
        rows = _brush_contribution_template_rows(db, brush_id=brush_id)
        return _file_response(
            resource,
            file_format,
            rows,
            purpose="template",
            include_metadata=True,
            fields=fields,
        )
    rows: list[dict[str, Any]] = []
    return _file_response(
        resource,
        file_format,
        rows,
        purpose="template",
        include_metadata=True,
        fields=_upload_template_fields(
            resource.fields,
            hidden_fields=set((default_values or {}).keys()),
        ),
    )


def export_resource(
    resource_key: str,
    file_format: FileFormat,
    db: Session,
    *,
    quality_type: str | None = None,
) -> Response:
    resource = get_resource(resource_key)
    if resource.key == "quality.measurements":
        fields = tuple(
            field
            for field in _quality_measurement_bulk_fields(quality_type)
            if field.name != "id"
        )
        rows = _quality_measurement_export_rows(db, quality_type=quality_type)
        return _file_response(
            resource,
            file_format,
            rows,
            purpose="export",
            include_metadata=file_format == "xlsx",
            fields=fields,
        )
    rows = [_resource_row(resource, db, item) for item in _query_resources(resource, db)]
    return _file_response(
        resource,
        file_format,
        rows,
        purpose="export",
        include_metadata=file_format == "xlsx",
        fields=_upload_template_fields(resource.fields),
    )


def import_resource(
    resource_key: str,
    content: bytes,
    *,
    filename: str,
    mode: ImportMode,
    default_values: dict[str, Any] | None = None,
    progress_callback: Callable[[], None] | None = None,
    db: Session,
) -> dict[str, Any]:
    resource = get_resource(resource_key)
    if not resource.importable:
        raise HTTPException(status_code=405, detail=f"{resource.label}不支持批量导入")
    if resource.key == "quality.measurements":
        return _import_quality_measurements(
            content,
            filename=filename,
            mode=mode,
            progress_callback=progress_callback,
            db=db,
        )
    if resource.key == "process.brush-contributions":
        return _import_brush_contributions(
            content,
            filename=filename,
            mode=mode,
            default_values=default_values,
            progress_callback=progress_callback,
            db=db,
        )
    rows = _parse_rows(content, filename, fields=resource.fields)
    field_names = {field.name for field in resource.fields}
    required = {field.name for field in resource.fields if field.required}
    default_values = default_values or {}
    unknown = sorted({key for row in rows for key in row if key and key not in field_names})
    if unknown:
        raise HTTPException(status_code=422, detail=f"模板字段不匹配，未知字段：{', '.join(unknown)}")
    unknown_defaults = sorted(key for key in default_values if key not in field_names)
    if unknown_defaults:
        raise HTTPException(
            status_code=422,
            detail=f"默认导入字段不匹配：{', '.join(unknown_defaults)}",
        )

    created = 0
    updated = 0
    skipped = 0
    errors: list[dict[str, Any]] = []
    reference_cache: dict[tuple[str, ...], str] = {}
    reference_candidate_cache: dict[
        tuple[type, tuple[str, ...], tuple[str, ...]], list[tuple[Any, set[str]]]
    ] = {}
    for row_index, raw_row in enumerate(rows, start=2):
        if progress_callback:
            progress_callback()
        if not any(str(value or "").strip() for value in raw_row.values()):
            skipped += 1
            continue
        merged_row = _apply_default_values(raw_row, default_values)
        try:
            resolved_row = _resolve_business_references(
                resource,
                db,
                merged_row,
                cache=reference_cache,
                candidate_cache=reference_candidate_cache,
            )
            normalized = _normalize_row(resource, resolved_row)
            missing = sorted(
                name
                for name in required
                if name != "id" and _is_blank(normalized.get(name))
            )
            if missing:
                labels = [
                    _display_field_label(next(field for field in resource.fields if field.name == name))
                    for name in missing
                ]
                raise ValueError(f"缺少必填内容：{', '.join(labels)}")
            existing_id = _resolve_existing_id(resource, db, normalized) if mode == "upsert" else None
            if existing_id:
                if not resource.update_func or not resource.update_schema:
                    skipped += 1
                    continue
                _call_update(resource, db, existing_id, normalized)
                updated += 1
            else:
                _call_create(resource, db, normalized)
                created += 1
        except Exception as exc:  # noqa: BLE001 - row-level import must collect all failures.
            db.rollback()
            errors.append({"row": row_index, "message": _error_message(exc)})

    return {
        "resource_key": resource.key,
        "resource_label": resource.label,
        "mode": mode,
        "total_rows": len(rows),
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "failed": len(errors),
        "errors": errors[:100],
        "truncated_errors": len(errors) > 100,
    }


def _build_fields(
    create_schema: type[BaseModel],
    update_schema: type[BaseModel] | None,
    extra_fields: tuple[BulkField, ...],
) -> tuple[BulkField, ...]:
    fields = [BulkField("id", "ID（可选，用于更新）", "string", False, "", "导出后回写更新使用")]
    fields.extend(extra_fields)
    seen = {field.name for field in fields}
    for schema in [create_schema, update_schema]:
        if not schema:
            continue
        for name, model_field in schema.model_fields.items():
            if name in seen:
                continue
            fields.append(
                BulkField(
                    name=name,
                    label=_human_field_label(name),
                    kind=_kind_from_annotation(model_field.annotation),
                    required=schema is create_schema and model_field.is_required(),
                    example=_example_for(name, model_field.annotation),
                    description=_description_for(name),
                )
            )
            seen.add(name)
    return tuple(fields)


def _kind_from_annotation(annotation: Any) -> str:
    annotation = _strip_optional(annotation)
    origin = get_origin(annotation)
    if annotation is list or origin is list:
        return "list"
    if annotation is dict or origin is dict:
        return "json"
    if annotation in {int}:
        return "integer"
    if annotation in {float}:
        return "number"
    if annotation is bool:
        return "boolean"
    if annotation is datetime:
        return "datetime"
    return "string"


def _strip_optional(annotation: Any) -> Any:
    origin = get_origin(annotation)
    if origin in {Union, UnionType}:
        args = [arg for arg in get_args(annotation) if arg is not type(None)]
        if len(args) == 1:
            return args[0]
    return annotation


def _example_for(name: str, annotation: Any) -> str:
    kind = _kind_from_annotation(annotation)
    if name.endswith("_id"):
        return "由页面范围自动带入，或填写业务编号"
    if kind == "boolean":
        return "是"
    if kind == "number":
        return "12.34"
    if kind == "integer":
        return "1"
    if kind == "datetime":
        return "2026-06-10 08:00:00"
    if kind == "json":
        return "项目=数值；项目=数值"
    if kind == "list":
        return "多个内容用逗号分隔"
    if "quality_type" in name or name == "target_family":
        return "橘皮"
    if "process_stage" in name:
        return "中涂外喷"
    return ""


def _description_for(name: str) -> str:
    descriptions = {
        "id": "系统内部记录号；新增模板不会要求填写。",
        "metrics": "质量指标请使用质量数据专用中文模板填写。",
        "repeat_readings": "逐次读数请使用仪器文件导入或逐行表格填写。",
        "vehicle_model_ids": "适用车型；多个内容用逗号分隔。",
        "color_ids": "适用颜色；多个内容用逗号分隔。",
        "supported_quality_types": "支持的质量类型；多个内容用逗号分隔。",
        "target_families": "批准的质量目标；多个内容用逗号分隔。",
    }
    if name.endswith("_id"):
        return "填写页面中看到的业务代码或名称；系统会自动查找对应记录，不需要查询内部编号。"
    if name.endswith("_ids"):
        return "填写页面中看到的业务代码或名称；多个内容用逗号分隔，系统会自动查找对应记录。"
    return descriptions.get(name, "")


_FIELD_LABEL_OVERRIDES = {
    "id": "系统记录号",
    "code": "业务代码",
    "name": "名称",
    "remark": "备注",
    "description": "说明",
    "factory_id": "工厂",
    "vehicle_model_id": "车型",
    "vehicle_model_ids": "适用车型",
    "color_id": "颜色",
    "color_ids": "适用颜色",
    "part_id": "零件",
    "measurement_group_id": "测量编组",
    "measurement_point_id": "测量点",
    "production_run_id": "生产记录",
    "production_stage_run_id": "生产工序记录",
    "program_version_id": "喷涂程序版本",
    "spray_program_id": "喷涂程序",
    "brush_id": "刷子号",
    "instrument_id": "测量仪器",
    "measurement_probe_id": "测量探头",
    "measurement_method_id": "测量方法",
    "calibration_record_id": "校准记录",
    "reference_standard_id": "参考件",
    "import_profile_id": "导入规则",
    "process_route_id": "工艺路线",
    "task_id": "问题任务",
    "material_batch_id": "材料批次",
    "characteristic_definition_id": "材料特性",
    "method_id": "检测方法",
    "robot_id": "机器人",
    "controller_id": "应用控制器",
    "atomizer_id": "静电旋杯",
    "path_segment_id": "轨迹段",
    "trajectory_program_id": "轨迹程序",
    "contribution_version_id": "贡献关系版本",
    "endpoint_id": "对接系统",
    "body_no": "生产车号",
    "run_no": "生产记录编号",
    "batch_no": "材料批次号",
    "process_stage": "喷涂工序",
    "quality_type": "质量指标类型",
    "quality_types": "适用质量类型",
    "is_match_point": "是否匹配点",
    "data_type": "数据用途",
    "measured_at": "检测时间",
    "measured_by": "检测人员",
    "started_at": "开始时间",
    "completed_at": "完成时间",
    "created_by": "创建人员",
    "approved_by": "批准人员",
    "status": "当前状态",
    "is_active": "是否启用",
    "is_valid": "数据是否有效",
    "is_approved": "是否已批准",
    "version": "版本号",
    "unit": "单位",
    "supplier": "供应商",
    "viscosity": "粘度",
    "solid_ratio": "固含量",
    "actual_value": "实际值",
    "configured_value": "设定值",
    "parameter_code": "参数代码",
    "parameter_name": "参数名称",
    "brush_no": "刷子号",
    "brush_table_no": "刷子表号",
    "spray_position": "喷涂位置",
    "overlap_ratio": "重叠率",
    "contribution_weight": "贡献权重",
    "data_no": "质量数据编号",
    "status_score": "状态评分",
    "site_owner": "现场调试负责人",
    "material_code": "材料代码",
    "material_name": "材料名称",
    "material_type": "材料类型",
    "process_immediately": "是否立即处理",
    "project_to_2d": "同步到二维点位图",
    "target_family": "目标质量类型",
    "target_families": "目标质量类型",
    "source_filename": "来源文件名",
    "error_report": "错误明细",
    "replay_of_job_id": "原导入任务",
    "target_resource": "导入到哪个业务模块",
    "field_mapping": "文件列对应关系",
    "required_fields": "必须填写的列",
    "validation_rules": "数据检查规则",
    "causality_status": "原因验证状态",
    "owner_user_id": "负责人账号",
    "problem_statement": "问题描述",
    "suspected_cause": "疑似原因",
    "data_quality_status": "数据检查状态",
    "material_status": "材料检查状态",
    "durr_execution_status": "杜尔设备执行检查状态",
    "symptom_pattern": "异常表现规律",
    "diagnosis_rule": "诊断规则",
    "recommended_checks": "建议检查项",
    "related_parameters": "相关工艺参数",
    "evidence_level": "证据等级",
    "grr_percent": "重复性与再现性占比",
    "raw_results": "原始结果明细",
    "geometry_class": "几何形状类别",
    "layer_scope": "适用涂层范围",
    "valid_until": "有效截止时间",
    "feature_impacts": "影响因素明细",
    "control_requirements": "管控要求",
    "bake_strategy": "烘烤节点说明",
    "containment_action": "临时控制措施",
    "supplier_response": "供应商回复",
    "deviation_decision": "偏差处置决定",
    "source_import_job_id": "来源导入任务",
    "start_position": "开始位置",
    "end_position": "结束位置",
    "normal_vector": "表面法向量",
    "gun_distance": "喷枪距离",
    "path_spacing": "轨迹间距",
    "collision_risk_score": "碰撞风险评分",
    "base_url": "服务地址",
    "max_attempts": "最多尝试次数",
    "canonical_unit": "标准单位",
    "result_unit": "结果单位",
    "supported_quality_types": "支持的质量类型",
    "minimum_repeats": "最少重复测量次数",
    "parameter_definition_id": "工艺参数",
    "is_recommendable": "是否允许系统建议调整",
    "actual_parameters": "实际工艺参数明细",
    "is_master_sample": "是否封样版本",
    "model_asset_key": "车身模型文件标识",
    "body_view": "车身视图",
    "grid_col": "网格列",
    "raw_file_uri": "原始文件位置",
    "bell_cup_type": "旋杯类型",
    "bell_cup_code": "旋杯代码",
    "deviation_details": "执行偏差明细",
    "configured_speed": "设定速度",
    "speed_unit": "速度单位",
    "trigger_state": "喷涂触发状态",
    "trigger_start_ms": "喷涂触发开始时间（毫秒）",
    "trigger_end_ms": "喷涂触发结束时间（毫秒）",
    "controller_software_version": "控制器软件版本",
}

_FIELD_TOKEN_LABELS = {
    "actual": "实际", "aggregation": "汇总", "ai": "智能分析", "approved": "批准",
    "auth": "认证", "author": "编写人", "bake": "烘烤", "base": "服务地址", "batch": "批次",
    "bell": "旋杯", "body": "车身", "brush": "刷子", "calibrated": "校准时间",
    "calibration": "校准", "canonical": "标准", "category": "类别", "causality": "原因验证",
    "certificate": "证书", "characteristic": "特性", "check": "检查", "checksum": "文件校验值",
    "coa": "批次检验报告", "coating": "涂层", "collision": "碰撞", "color": "颜色",
    "comment": "协作记录", "conditions": "适用条件", "confidence": "可信度", "config": "配置",
    "configuration": "配置", "constraint": "约束", "containment": "临时控制措施", "context": "生产上下文",
    "control": "管控", "controlled": "受控", "coordinate": "坐标", "data": "数据", "deviation": "偏差处置",
    "device": "设备", "diagnosis": "诊断", "digital": "数字标准", "direction": "方向",
    "document": "文件", "doe": "试验设计文件", "domain": "业务类别", "downstream": "下游",
    "due": "要求完成时间", "durr": "杜尔设备", "effective": "生效", "end": "结束",
    "endpoint": "对接系统", "entry": "知识条目", "error": "错误报告", "event": "事件",
    "evidence": "证据", "executed": "实际执行", "expected": "应有", "explanation": "解释",
    "factory": "工厂", "failed": "失败", "feature": "影响因素", "field": "字段",
    "firmware": "固件", "generated": "生成", "geometry": "几何", "grid": "网格",
    "grr": "重复性与再现性", "gun": "喷枪", "hard": "强制", "hypothesis": "原因假设",
    "import": "导入", "imported": "导入", "instructions": "操作说明", "instrument": "仪器", "issue": "问题",
    "layer": "涂层", "layout": "位置", "limit": "限值", "lower": "下限", "manufacturer": "制造商",
    "material": "材料", "max": "最大", "measured": "检测", "measurement": "测量",
    "method": "方法", "metric": "指标", "min": "最小", "minimum": "最少", "model": "模型",
    "msds": "安全说明文件", "ndc": "非离散度", "normal": "法向", "operator": "操作人员",
    "orientation": "姿态", "owner": "负责人", "parameter": "参数", "parser": "解析方式",
    "part": "零件", "path": "轨迹", "payload": "业务内容", "performed": "执行人员",
    "point": "点位", "pos": "位置", "prediction": "预测", "preview": "预览",
    "probe": "探头", "problem": "问题描述", "procedure": "作业文件", "process": "工艺",
    "production": "生产", "profile": "导入规则", "program": "程序", "quality": "质量",
    "raw": "原始", "recommendation": "建议", "recommended": "建议", "reference": "参考件",
    "region": "区域", "related": "相关", "repeat": "重复", "replay": "重放来源",
    "trial": "试验", "conclusion": "结论",
    "required": "必填", "requires": "是否需要", "resolution": "解决方案", "result": "结果",
    "reviewed": "复核", "robot": "机器人", "role": "角色", "route": "路线", "row": "行数",
    "sample": "样本", "sampled": "采集", "schema": "文件格式版本", "segment": "轨迹段", "sensitivity": "敏感性",
    "sequence": "顺序", "serial": "序列号", "severity": "严重程度", "shift": "班次",
    "site": "现场", "soft": "建议", "software": "软件", "solid": "固含量", "source": "来源",
    "speed": "速度", "spray": "喷涂", "standard": "标准", "start": "开始", "station": "工位",
    "step": "步骤", "study": "研究", "submission": "供应商提交", "submitted": "提交",
    "substrate": "基材", "summary": "摘要", "supplier": "供应商", "supported": "支持",
    "suspected": "疑似", "symptom": "异常表现", "system": "系统", "tags": "标签",
    "target": "目标", "task": "任务", "tcp": "工具中心点", "tds": "材料技术文件",
    "tested": "检测", "title": "标题", "trajectory": "轨迹", "trigger": "喷涂触发",
    "uncertainty": "不确定度", "upper": "上限", "upstream": "上游", "valid": "有效",
    "validation": "验证", "vehicle": "车型", "version": "版本", "viscosity": "粘度",
    "x": "横向", "y": "纵向", "z": "高度", "no": "编号", "type": "类型",
    "value": "值", "values": "数值", "uri": "文件位置", "score": "评分", "count": "数量",
    "time": "时间", "date": "日期", "at": "时间", "from": "开始", "to": "结束", "by": "人员",
    "is": "是否", "id": "", "ids": "", "code": "代码", "name": "名称",
}


def _human_field_label(name: str) -> str:
    if name in _FIELD_LABEL_OVERRIDES:
        return _FIELD_LABEL_OVERRIDES[name]
    translated = "".join(_FIELD_TOKEN_LABELS.get(token, token.upper()) for token in name.split("_"))
    return translated or "业务信息"


def _display_field_label(field_spec: BulkField) -> str:
    configured = field_spec.label.strip()
    if configured and configured != field_spec.name:
        return (
            configured.replace(" ID", "")
            .replace("id ", "")
            .replace("id时", "记录号时")
        )
    return _human_field_label(field_spec.name)


def _upload_template_fields(
    fields: tuple[BulkField, ...],
    *,
    hidden_fields: set[str] | None = None,
) -> tuple[BulkField, ...]:
    hidden = {"id", *(hidden_fields or set())}
    return tuple(field for field in fields if field.name not in hidden)


_REFERENCE_CONTEXT_FIELDS = (
    "factory_id",
    "vehicle_model_id",
    "color_id",
    "part_id",
    "spray_program_id",
    "program_version_id",
    "instrument_id",
    "characteristic_definition_id",
    "process_route_id",
    "production_run_id",
    "trajectory_program_id",
)


def _normalize_reference_text(value: Any) -> str:
    return re.sub(r"\s*[/|:]\s*", "/", str(value or "").strip()).casefold()


def _business_reference_item_display(db: Session, reference: BusinessReference, item: Any) -> str:
    if isinstance(item, SprayProgramVersion):
        program = db.get(SprayProgram, item.spray_program_id)
        return " / ".join(value for value in [getattr(program, "program_code", ""), item.version] if value)
    if isinstance(item, Brush):
        version = db.get(SprayProgramVersion, item.program_version_id)
        program = db.get(SprayProgram, version.spray_program_id) if version else None
        return " / ".join(
            value
            for value in [getattr(program, "program_code", ""), getattr(version, "version", ""), item.brush_no]
            if value
        )
    if isinstance(item, ProductionStageRun):
        production_run = db.get(ProductionRun, item.production_run_id)
        return " / ".join(value for value in [getattr(production_run, "run_no", ""), item.process_stage] if value)
    if isinstance(item, TrajectoryPathSegment):
        trajectory = db.get(TrajectoryProgram, item.trajectory_program_id)
        return " / ".join(
            str(value)
            for value in [getattr(trajectory, "trajectory_code", ""), getattr(trajectory, "version", ""), item.segment_no]
            if value not in {None, ""}
        )
    values = [getattr(item, field_name, None) for field_name in reference.display_fields]
    return " / ".join(str(value) for value in values if value not in {None, ""})


def _business_reference_display(db: Session, reference: BusinessReference, value: Any) -> str:
    item = db.get(reference.model, str(value))
    if not item:
        return str(value)
    return _business_reference_item_display(db, reference, item) or str(value)


def _reference_candidate_tokens(db: Session, reference: BusinessReference, item: Any) -> set[str]:
    values = {
        _normalize_reference_text(getattr(item, field_name, ""))
        for field_name in reference.lookup_fields
        if getattr(item, field_name, None) not in {None, ""}
    }
    display = _business_reference_item_display(db, reference, item)
    if display:
        values.add(_normalize_reference_text(display))
    return values


def _resolve_business_reference_value(
    resource: BulkResource,
    db: Session,
    field_name: str,
    value: Any,
    row: dict[str, Any],
    *,
    cache: dict[tuple[str, ...], str],
    candidate_cache: dict[
        tuple[type, tuple[str, ...], tuple[str, ...]], list[tuple[Any, set[str]]]
    ],
) -> str:
    reference = _business_reference_for(resource, field_name)
    if not reference:
        return str(value)
    value_text = str(value).strip()
    existing = db.get(reference.model, value_text)
    if existing:
        return str(existing.id)

    context_values = tuple(
        str(row.get(context_field) or "")
        for context_field in _REFERENCE_CONTEXT_FIELDS
        if context_field != field_name
    )
    cache_key = (resource.key, field_name, _normalize_reference_text(value_text), *context_values)
    if cache_key in cache:
        return cache[cache_key]

    candidate_cache_key = (
        reference.model,
        reference.lookup_fields,
        reference.display_fields,
    )
    candidate_entries = candidate_cache.get(candidate_cache_key)
    if candidate_entries is None:
        candidate_entries = [
            (item, _reference_candidate_tokens(db, reference, item))
            for item in db.scalars(select(reference.model))
        ]
        candidate_cache[candidate_cache_key] = candidate_entries
    for context_field in _REFERENCE_CONTEXT_FIELDS:
        context_value = row.get(context_field)
        if context_field == field_name or _is_blank(context_value):
            continue
        if not hasattr(reference.model, context_field):
            continue
        narrowed = [
            entry
            for entry in candidate_entries
            if str(getattr(entry[0], context_field, "")) == str(context_value)
        ]
        if narrowed:
            candidate_entries = narrowed

    normalized_value = _normalize_reference_text(value_text)
    matches = [
        item
        for item, tokens in candidate_entries
        if normalized_value in tokens
    ]
    if not matches:
        raise ValueError(f"{reference.label}“{value_text}”不存在，请检查业务代码或名称")
    if len(matches) > 1:
        examples = "、".join(
            _business_reference_item_display(db, reference, item)
            for item in matches[:3]
        )
        raise ValueError(
            f"{reference.label}“{value_text}”对应多条记录，请填写完整的代码和版本，例如：{examples}"
        )
    resolved_id = str(matches[0].id)
    cache[cache_key] = resolved_id
    return resolved_id


def _resolve_business_references(
    resource: BulkResource,
    db: Session,
    row: dict[str, Any],
    *,
    cache: dict[tuple[str, ...], str],
    candidate_cache: dict[
        tuple[type, tuple[str, ...], tuple[str, ...]], list[tuple[Any, set[str]]]
    ],
) -> dict[str, Any]:
    resolved = dict(row)
    ordered_fields = list(_REFERENCE_CONTEXT_FIELDS) + [
        field_name for field_name in resolved if field_name not in _REFERENCE_CONTEXT_FIELDS
    ]
    for field_name in ordered_fields:
        if field_name not in resolved or _is_blank(resolved[field_name]):
            continue
        singular_name = {
            "vehicle_model_ids": "vehicle_model_id",
            "color_ids": "color_id",
        }.get(field_name, field_name)
        reference = _business_reference_for(resource, singular_name)
        if not reference:
            continue
        if field_name.endswith("_ids"):
            resolved[field_name] = [
                _resolve_business_reference_value(
                    resource,
                    db,
                    singular_name,
                    item,
                    resolved,
                    cache=cache,
                    candidate_cache=candidate_cache,
                )
                for item in _coerce_list(resolved[field_name])
            ]
        else:
            resolved[field_name] = _resolve_business_reference_value(
                resource,
                db,
                field_name,
                resolved[field_name],
                resolved,
                cache=cache,
                candidate_cache=candidate_cache,
            )
    return resolved


def _query_resources(resource: BulkResource, db: Session) -> list[Any]:
    query = select(resource.model)
    order_columns = [
        getattr(resource.model, name)
        for name in resource.order_by
        if hasattr(resource.model, name)
    ]
    if order_columns:
        query = query.order_by(*order_columns)
    return list(db.scalars(query))


def _resource_row(resource: BulkResource, db: Session, item: Any) -> dict[str, Any]:
    if resource.export_row:
        row = resource.export_row(db, item)
    else:
        row = {field.name: getattr(item, field.name, None) for field in resource.fields}
    row.setdefault("id", getattr(item, "id", ""))
    exported: dict[str, Any] = {}
    for field_spec in resource.fields:
        value = row.get(field_spec.name)
        if field_spec.name in {"vehicle_model_ids", "color_ids"} and isinstance(value, list):
            reference = _BUSINESS_REFERENCES[
                "vehicle_model_id" if field_spec.name == "vehicle_model_ids" else "color_id"
            ]
            value = [
                _business_reference_display(db, reference, item_id)
                for item_id in value
            ]
        else:
            reference = _business_reference_for(resource, field_spec.name)
            if reference and not _is_blank(value):
                value = _business_reference_display(db, reference, value)
        exported[field_spec.name] = value
    return exported


def _file_response(
    resource: BulkResource,
    file_format: FileFormat,
    rows: list[dict[str, Any]],
    *,
    purpose: Literal["template", "export"],
    include_metadata: bool,
    fields: tuple[BulkField, ...] | None = None,
) -> Response:
    field_specs = fields or resource.fields
    if file_format == "xlsx":
        content = _xlsx_bytes(field_specs, rows, include_metadata=include_metadata)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        content = _csv_bytes(field_specs, rows)
        media_type = "text/csv; charset=utf-8"
    suffix = "xlsx" if file_format == "xlsx" else "csv"
    filename = f"{resource.key}-{purpose}.{suffix}"
    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _csv_bytes(fields: tuple[BulkField, ...], rows: list[dict[str, Any]]) -> bytes:
    output = StringIO()
    headers = [_display_field_label(field) for field in fields]
    _ensure_unique_display_labels(fields, headers)
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    for row in rows:
        writer.writerow(
            {
                _display_field_label(field): _cell_to_string(row.get(field.name))
                for field in fields
            }
        )
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def _xlsx_bytes(
    fields: tuple[BulkField, ...],
    rows: list[dict[str, Any]],
    *,
    include_metadata: bool,
) -> bytes:
    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "数据"
    headers = [_display_field_label(field) for field in fields]
    _ensure_unique_display_labels(fields, headers)
    data_sheet.append(headers)
    for row in rows:
        data_sheet.append([_cell_to_string(row.get(field.name)) for field in fields])
    data_sheet.freeze_panes = "A2"
    for column in data_sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        data_sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 60)

    if include_metadata:
        meta = workbook.create_sheet("填写说明")
        meta.append(["列名", "填写类型", "是否必填", "填写示例", "填写说明"])
        for field_spec in fields:
            meta.append(
                [
                    _display_field_label(field_spec),
                    _kind_label(field_spec.kind),
                    "是" if field_spec.required else "否",
                    field_spec.example,
                    field_spec.description,
                ]
            )
        meta.freeze_panes = "A2"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _parse_rows(
    content: bytes,
    filename: str,
    *,
    fields: tuple[BulkField, ...],
) -> list[dict[str, Any]]:
    if not content:
        raise HTTPException(status_code=422, detail="导入文件为空")
    lowered = filename.lower()
    if lowered.endswith(".xlsx"):
        rows = _parse_xlsx(content)
    else:
        rows = _parse_csv(content)
    return _canonicalize_input_rows(rows, fields)


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(status_code=422, detail="CSV 缺少表头")
    return [dict(row) for row in reader]


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    if "数据" in workbook.sheetnames:
        sheet = workbook["数据"]
    elif "data" in workbook.sheetnames:
        sheet = workbook["data"]
    else:
        sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=422, detail="Excel 缺少“数据”工作表表头")
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    if not any(headers):
        raise HTTPException(status_code=422, detail="Excel 表头为空")
    parsed = []
    for values in rows[1:]:
        parsed.append({headers[index]: values[index] if index < len(values) else None for index in range(len(headers)) if headers[index]})
    return parsed


def _kind_label(kind: str) -> str:
    return {
        "boolean": "是/否",
        "integer": "整数",
        "number": "数值",
        "datetime": "日期时间",
        "json": "分项内容",
        "list": "多个内容",
        "string": "文字",
    }.get(kind, "文字")


def _ensure_unique_display_labels(fields: tuple[BulkField, ...], headers: list[str]) -> None:
    duplicates = sorted({header for header in headers if headers.count(header) > 1})
    if duplicates:
        names = [field.name for field in fields if _display_field_label(field) in duplicates]
        raise ValueError(f"批量文件中文列名重复，请修正字段配置：{', '.join(names)}")


def _canonicalize_input_rows(
    rows: list[dict[str, Any]],
    fields: tuple[BulkField, ...],
) -> list[dict[str, Any]]:
    aliases: dict[str, str] = {}
    for field_spec in fields:
        aliases[field_spec.name] = field_spec.name
        aliases[_display_field_label(field_spec)] = field_spec.name
    normalized_rows: list[dict[str, Any]] = []
    for row in rows:
        normalized: dict[str, Any] = {}
        for raw_header, value in row.items():
            header = str(raw_header or "").strip()
            canonical = aliases.get(header, header)
            if canonical in normalized and not _is_blank(value):
                raise HTTPException(status_code=422, detail=f"文件中重复填写了列：{header}")
            normalized[canonical] = value
        normalized_rows.append(normalized)
    return normalized_rows


def _normalize_row(resource: BulkResource, raw_row: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    fields = {field.name: field for field in resource.fields}
    for name, field_spec in fields.items():
        if name not in raw_row:
            continue
        normalized[name] = _coerce_value(raw_row[name], field_spec)
    return normalized


_BUSINESS_VALUE_ALIASES: dict[str, dict[str, str]] = {
    "quality_type": {"橘皮": "ORANGE_PEEL", "色差": "COLOR_DIFFERENCE", "膜厚": "THICKNESS"},
    "quality_types": {"橘皮": "ORANGE_PEEL", "色差": "COLOR_DIFFERENCE", "膜厚": "THICKNESS"},
    "supported_quality_types": {"橘皮": "ORANGE_PEEL", "色差": "COLOR_DIFFERENCE", "膜厚": "THICKNESS"},
    "target_quality_type": {"橘皮": "ORANGE_PEEL", "色差": "COLOR_DIFFERENCE", "膜厚": "THICKNESS"},
    "target_family": {"橘皮": "ORANGE_PEEL", "色差": "COLOR_DIFFERENCE", "膜厚": "THICKNESS"},
    "target_families": {"橘皮": "ORANGE_PEEL", "色差": "COLOR_DIFFERENCE", "膜厚": "THICKNESS"},
    "process_stage": {
        "中涂外喷": "MIDCOAT_EXT",
        "色漆一站": "BASECOAT_1",
        "色漆二站": "BASECOAT_2",
        "清漆一站": "CLEARCOAT_1",
        "清漆二站": "CLEARCOAT_2",
    },
    "material_type": {"中涂": "MIDCOAT", "色漆": "BASECOAT", "清漆": "CLEARCOAT"},
    "coating_system": {"中涂": "MIDCOAT", "色漆": "BASECOAT", "清漆": "CLEARCOAT"},
    "color_type": {"中涂颜色": "MIDCOAT", "色漆颜色": "BASECOAT"},
    "data_type": {"测试数据": "TEST", "封样数据": "MASTER_SAMPLE", "标准数据": "STANDARD"},
    "point_type": {"质量检测": "QUALITY", "工艺检查": "PROCESS", "材料检查": "MATERIAL"},
    "measurement_direction": {"纵向": "LONGITUDINAL", "横向": "TRANSVERSE", "多方向": "MULTI_DIRECTION"},
    "source": {"专家评估": "EXPERT", "设备仿真": "SIMULATION", "现场试验": "DOE", "沉积模型": "DEPOSITION_MODEL"},
}


def _translate_business_value(field_name: str, value: Any) -> Any:
    aliases = _BUSINESS_VALUE_ALIASES.get(field_name)
    if not aliases or not isinstance(value, str):
        return value
    return aliases.get(value.strip(), value)


def _apply_default_values(
    raw_row: dict[str, Any],
    default_values: dict[str, Any],
) -> dict[str, Any]:
    merged = dict(raw_row)
    for key, value in default_values.items():
        if key not in merged or _is_blank(merged.get(key)):
            merged[key] = value
    return merged


def _coerce_value(value: Any, field_spec: BulkField) -> Any:
    if _is_blank(value):
        return None
    if isinstance(value, str):
        value = value.strip()
    if field_spec.kind == "boolean":
        return _coerce_bool(value)
    if field_spec.kind == "integer":
        return int(value)
    if field_spec.kind == "number":
        return float(value)
    if field_spec.kind == "datetime":
        return _coerce_datetime(value)
    if field_spec.kind == "json":
        return _coerce_json(value, object_required=True)
    if field_spec.kind == "list":
        return [
            _translate_business_value(field_spec.name, item)
            for item in _coerce_list(value)
        ]
    value = _translate_business_value(field_spec.name, value)
    return str(value) if not isinstance(value, str) else value


def _coerce_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in {"true", "1", "yes", "y", "是", "启用", "生效"}:
        return True
    if normalized in {"false", "0", "no", "n", "否", "停用", "失效"}:
        return False
    raise ValueError(f"布尔字段无法识别：{value}")


def _coerce_datetime(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    return datetime.fromisoformat(str(value).replace("Z", "+00:00"))


def _coerce_json(value: Any, *, object_required: bool) -> Any:
    if isinstance(value, (dict, list)):
        parsed = value
    else:
        text = str(value).strip()
        try:
            parsed = json.loads(text)
        except json.JSONDecodeError:
            pairs = [part.strip() for part in re.split(r"[;；]", text) if part.strip()]
            if not pairs or any("=" not in pair for pair in pairs):
                raise ValueError("分项内容格式不正确，请按“项目=值；项目=值”填写") from None
            parsed = {
                key.strip(): _infer_business_scalar(item_value.strip())
                for key, item_value in (pair.split("=", 1) for pair in pairs)
                if key.strip()
            }
    if object_required and not isinstance(parsed, dict):
        raise ValueError("分项内容格式不正确，请按模板逐项填写")
    return parsed


def _infer_business_scalar(value: str) -> Any:
    if value in {"是", "true", "True"}:
        return True
    if value in {"否", "false", "False"}:
        return False
    try:
        numeric = float(value)
        return int(numeric) if numeric.is_integer() else numeric
    except ValueError:
        return value


def _coerce_list(value: Any) -> list[Any]:
    if isinstance(value, list):
        return value
    text = str(value).strip()
    if text.startswith("["):
        parsed = json.loads(text)
        if not isinstance(parsed, list):
            raise ValueError("多个内容请使用逗号分隔")
        return parsed
    return [item.strip() for item in re.split(r"[,，]", text) if item.strip()]


def _cell_to_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, dict):
        text = "；".join(
            f"{key}={_human_cell_value(item_value)}"
            for key, item_value in value.items()
        )
    elif isinstance(value, list):
        text = "，".join(_human_cell_value(item) for item in value)
    elif isinstance(value, bool):
        text = "是" if value else "否"
    else:
        text = str(value)
    return _escape_spreadsheet_text(text)


def _human_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        text = "是" if value else "否"
    elif isinstance(value, dict):
        text = "；".join(f"{key}={_human_cell_value(item)}" for key, item in value.items())
    elif isinstance(value, list):
        text = "，".join(_human_cell_value(item) for item in value)
    else:
        text = str(value)
    return _escape_spreadsheet_text(text)


def _escape_spreadsheet_text(value: str) -> str:
    if value.lstrip(" \t\r\n").startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _resolve_existing_id(resource: BulkResource, db: Session, row: dict[str, Any]) -> str | None:
    row_id = row.get("id")
    if row_id and db.get(resource.model, row_id):
        return str(row_id)
    if resource.match_existing:
        return resource.match_existing(db, row)
    if not resource.unique_fields or any(_is_blank(row.get(field_name)) for field_name in resource.unique_fields):
        return None
    conditions = [getattr(resource.model, field_name) == row[field_name] for field_name in resource.unique_fields]
    existing = db.scalar(select(resource.model).where(*conditions))
    return existing.id if existing else None


def _call_create(resource: BulkResource, db: Session, row: dict[str, Any]) -> Any:
    payload = resource.create_schema(**_schema_values(resource.create_schema, row))
    args = [_required_arg(row, field_name) for field_name in resource.create_arg_fields]
    if args:
        return resource.create_func(*args, payload, db)
    return resource.create_func(payload, db)


def _call_update(resource: BulkResource, db: Session, resource_id: str, row: dict[str, Any]) -> Any:
    if not resource.update_func or not resource.update_schema:
        raise ValueError(f"{resource.label}不支持批量更新")
    payload = resource.update_schema(**_schema_values(resource.update_schema, row, exclude_none=False))
    if resource.update_uses_resource_id:
        return resource.update_func(resource_id, payload, db)
    args = [_required_arg(row, field_name) for field_name in resource.update_arg_fields]
    return resource.update_func(*args, payload, db)


def _schema_values(
    schema: type[BaseModel],
    row: dict[str, Any],
    *,
    exclude_none: bool = True,
) -> dict[str, Any]:
    values = {}
    for name in schema.model_fields:
        if name not in row:
            continue
        if exclude_none and row[name] is None:
            continue
        values[name] = row[name]
    return values


def _required_arg(row: dict[str, Any], field_name: str) -> Any:
    value = row.get(field_name)
    if _is_blank(value):
        raise ValueError(f"缺少上下文字段：{field_name}")
    return value


def _error_message(exc: Exception) -> str:
    if isinstance(exc, HTTPException):
        return str(exc.detail)
    if isinstance(exc, ValidationError):
        return "; ".join(
            f"{'.'.join(str(part) for part in error['loc'])}: {error['msg']}"
            for error in exc.errors()
        )
    return str(exc)


def _program_version_row(db: Session, version: SprayProgramVersion) -> dict[str, Any]:
    return {
        "id": version.id,
        "spray_program_id": version.spray_program_id,
        "version": version.version,
        "status": version.status,
        "source_type": version.source_type,
        "is_master_sample": version.is_master_sample,
        "approved_by": version.approved_by,
        "vehicle_model_ids": list(
            db.scalars(
                select(ProgramVehicleModel.vehicle_model_id).where(
                    ProgramVehicleModel.program_version_id == version.id
                )
            )
        ),
        "color_ids": list(
            db.scalars(
                select(ProgramColor.color_id).where(ProgramColor.program_version_id == version.id)
            )
        ),
    }


def _quality_measurement_row(db: Session, measurement: QualityMeasurement) -> dict[str, Any]:
    return {
        "id": measurement.id,
        "data_no": measurement.data_no,
        "production_run_id": measurement.production_run_id,
        "measurement_group_id": measurement.measurement_group_id,
        "measurement_point_id": measurement.measurement_point_id,
        "quality_type": measurement.quality_type,
        "data_type": measurement.data_type,
        "measured_at": measurement.measured_at,
        "measured_by": measurement.measured_by,
        "device_code": measurement.device_code,
        "instrument_id": measurement.instrument_id,
        "measurement_probe_id": measurement.measurement_probe_id,
        "measurement_method_id": measurement.measurement_method_id,
        "calibration_record_id": measurement.calibration_record_id,
        "reference_standard_id": measurement.reference_standard_id,
        "import_profile_id": measurement.import_profile_id,
        "measurement_direction": measurement.measurement_direction,
        "raw_file_uri": measurement.raw_file_uri,
        "status_score": measurement.status_score,
        "is_valid": measurement.is_valid,
        "metrics": [
            {
                "metric_code": metric.metric_code,
                "metric_name": metric.metric_name,
                "raw_value": metric.raw_value,
                "corrected_value": metric.corrected_value,
                "unit": metric.unit,
            }
            for metric in db.scalars(
                select(QualityMetricValue)
                .where(QualityMetricValue.measurement_id == measurement.id)
                .order_by(QualityMetricValue.metric_code)
            )
        ],
        "repeat_readings": [
            {
                "repeat_no": reading.repeat_no,
                "metric_code": reading.metric_code,
                "raw_value": reading.raw_value,
                "corrected_value": reading.corrected_value,
                "unit": reading.unit,
                "is_valid": reading.is_valid,
                "invalid_reason": reading.invalid_reason,
            }
            for reading in db.scalars(
                select(MeasurementRepeatReading)
                .where(MeasurementRepeatReading.measurement_id == measurement.id)
                .order_by(MeasurementRepeatReading.repeat_no, MeasurementRepeatReading.metric_code)
            )
        ],
    }


def _quality_metric_field_name(metric_code: str) -> str:
    return f"metric__{metric_code}"


def _quality_metric_code_from_field(field_name: str) -> str | None:
    if not field_name.startswith("metric__"):
        return None
    return field_name.split("__", 1)[1].strip()


def _quality_metric_catalog_for(quality_type: str | None = None) -> list[dict[str, Any]]:
    normalized = quality_type.strip().upper() if quality_type else None
    catalog = [
        metric for metric in QUALITY_METRIC_CATALOG if not normalized or metric["quality_type"] == normalized
    ]
    if normalized and not catalog:
        raise HTTPException(status_code=422, detail=f"不支持的质量类型：{normalized}")
    return catalog


_QUALITY_TYPE_DATA_NO_CODE = {
    "ORANGE_PEEL": "OP",
    "COLOR_DIFFERENCE": "CD",
    "THICKNESS": "TH",
}


def _sanitize_data_no_token(value: str, *, fallback: str) -> str:
    cleaned = re.sub(r"[^\w\-]+", "-", str(value or "").strip(), flags=re.UNICODE)
    cleaned = re.sub(r"-{2,}", "-", cleaned).strip("-_")
    return cleaned or fallback


def _build_quality_data_no(*, body_no: str, point_code: str, quality_type: str) -> str:
    body = _sanitize_data_no_token(body_no, fallback="BODY")
    point = _sanitize_data_no_token(point_code, fallback="POINT")
    quality = str(quality_type or "").strip().upper()
    quality_code = _QUALITY_TYPE_DATA_NO_CODE.get(quality, _sanitize_data_no_token(quality, fallback="QT"))
    return f"QM-{body}-{point}-{quality_code}"


def _quality_measurement_base_fields() -> tuple[BulkField, ...]:
    return (
        BulkField("id", "ID（可选，用于更新）", "string", False, "", "存在 id 时可直接更新对应质量测量。"),
        BulkField(
            "data_no",
            "质量数据编号",
            "string",
            False,
            "",
            "可留空。留空时后端按「车号 + 测量点代码 + 质量类型」自动生成，例如 QM-BODY-0001-PT1-OP。手工填写时仍可用于覆盖。",
        ),
        BulkField("production_run_no", "生产事件编号", "string", False, "RT_RUN_001", "可留空；留空时按车号自动生成 RUN-车号。若该生产事件不存在，后端会结合同一行的车号与生产上下文自动补建。"),
        BulkField("body_no", "车号", "string", False, "BODY-0001", "建议始终填写车号；生产事件编号留空或尚不存在时，车号用于自动创建生产事件，并参与自动生成 data_no。"),
        BulkField("factory_code", "工厂代码", "string", False, "F01", "当生产事件尚不存在时必填，用于自动创建生产事件。"),
        BulkField("measurement_group_code", "测量编组代码", "string", False, "RT_MG_OP", "模板会自动预填当前编组代码，导入时按车型+代码解析。"),
        BulkField("measurement_group_name", "测量编组名称", "string", False, "", "模板说明列，仅用于帮助识别编组，可保留原值。"),
        BulkField("measurement_point_code", "测量点代码", "string", True, "RT_P01", "模板会自动预填当前编组内测量点代码，导入时按车型+代码解析。"),
        BulkField("measurement_point_name", "测量点名称", "string", False, "", "模板说明列，仅用于帮助识别测量点，可保留原值。"),
        BulkField("vehicle_model_code", "车型代码", "string", False, "", "模板会自动预填当前测量点所属车型；当生产事件尚不存在时会参与自动创建。"),
        BulkField("vehicle_model_name", "车型名称", "string", False, "", "模板说明列，帮助用户确认当前测量点所属车型。"),
        BulkField("part_code", "零件代码", "string", False, "", "模板说明列，帮助用户确认当前测量点所属零件。"),
        BulkField("part_name", "零件名称", "string", False, "", "模板说明列，帮助用户确认当前测量点所属零件。"),
        BulkField("quality_type", "质量类型", "string", True, "橘皮", "可填写橘皮、色差或膜厚；模板会根据当前页面自动预填。"),
        BulkField("color_code", "颜色代码", "string", False, "C01", "当生产事件尚不存在时必填，用于自动创建生产事件。"),
        BulkField("shift", "班次", "string", False, "A", "可选；自动创建生产事件时会一并写入。"),
        BulkField("production_started_at", "生产开始时间", "datetime", False, "2026-07-08T07:30:00+08:00", "当生产事件尚不存在时建议填写；留空时默认回落到测量时间。"),
        BulkField("production_completed_at", "生产结束时间", "datetime", False, "2026-07-08T08:10:00+08:00", "可选；自动创建生产事件时会一并写入。"),
        BulkField("measured_at", "测量时间", "datetime", True, "2026-07-08T08:00:00+08:00", "支持 Excel 日期单元格或 ISO 时间。"),
        BulkField("measured_by", "测量人", "string", False, "张三", "可选。"),
        BulkField("data_type", "数据用途", "string", False, "测试数据", "可填写测试数据、封样数据或标准数据；未填写时按测试数据处理。"),
        BulkField("device_code", "设备编码", "string", False, "BYK-WAVE-01", "可选，未填写时如绑定仪器会自动带出仪器编码。"),
        BulkField("instrument_code", "仪器编码", "string", False, "BYK-WAVE-01", "按仪器编码解析，不要求用户手填 instrument_id。"),
        BulkField("measurement_method_code", "测量方法编码", "string", False, "BYK-WAVE-DOI:1", "可填写 code 或 code:version；若同编码存在多版本，必须带版本。"),
        BulkField("calibration_no", "校准记录编号", "string", False, "CAL-OP-1", "按校准记录编号解析。"),
        BulkField("reference_standard_code", "参考件编码", "string", False, "REF-OP-1", "按参考件编码解析。"),
        BulkField("import_profile_code", "导入模板编码", "string", False, "BYK-IMPORT:1", "可填写 code 或 code:version；若同编码存在多版本，必须带版本。"),
        BulkField("measurement_direction", "测量方向", "string", False, "纵向", "可填写纵向、横向或多方向。"),
        BulkField("raw_file_uri", "原始文件位置", "string", False, "", "可选，用于追溯仪器原始文件。"),
        BulkField("status_score", "状态分数", "number", False, "98.6", "可选。"),
        BulkField("is_valid", "数据有效", "boolean", False, "true", "默认 true。"),
    )


def _quality_measurement_bulk_fields(quality_type: str | None = None) -> tuple[BulkField, ...]:
    quality_labels = {
        "ORANGE_PEEL": "橘皮指标",
        "COLOR_DIFFERENCE": "色差指标",
        "THICKNESS": "膜厚指标",
    }
    metric_fields = tuple(
        BulkField(
            _quality_metric_field_name(metric["code"]),
            f"{quality_labels.get(metric['quality_type'], '质量指标')}：{metric['name']}（{metric['code']}）",
            "number",
            False,
            "82.5",
            f"用户直接填写检测数值，系统会自动归入对应质量指标。单位：{metric.get('unit') or '无'}",
        )
        for metric in _quality_metric_catalog_for(quality_type)
    )
    return _quality_measurement_base_fields() + metric_fields


def _quality_measurement_template_fields(quality_type: str | None = None) -> tuple[BulkField, ...]:
    """Upload template omits id/data_no — data_no is generated on import."""
    return tuple(
        field
        for field in _quality_measurement_bulk_fields(quality_type)
        if field.name not in {"id", "data_no"}
    )


def _normalize_optional_text(value: Any) -> str | None:
    if _is_blank(value):
        return None
    text = str(value).strip()
    return text or None


def _resolve_or_create_quality_production_run(
    *,
    db: Session,
    runs_by_no: dict[str, ProductionRun],
    factories_by_code: dict[str, Factory],
    vehicle_models_by_code: dict[str, VehicleModel],
    colors_by_code: dict[str, Color],
    row: dict[str, Any],
    measured_at_value: Any,
) -> ProductionRun:
    run_no = str(row.get("production_run_no") or "").strip()
    body_no = _normalize_optional_text(row.get("body_no"))
    factory_code = _normalize_optional_text(row.get("factory_code"))
    vehicle_model_code = _normalize_optional_text(row.get("vehicle_model_code"))
    color_code = _normalize_optional_text(row.get("color_code"))
    shift = _normalize_optional_text(row.get("shift"))
    started_at_value = row.get("production_started_at")
    completed_at_value = row.get("production_completed_at")

    if not run_no:
        if not body_no:
            raise ValueError("缺少生产事件编号时，必须填写车号以便自动生成生产事件编号")
        run_no = f"RUN-{body_no}"

    run = runs_by_no.get(run_no)
    if run:
        if body_no and run.body_no and run.body_no != body_no:
            raise ValueError(f"生产事件 {run_no} 的车号与导入内容不一致")
        if factory_code:
            factory = factories_by_code.get(factory_code)
            if not factory:
                raise ValueError(f"工厂 {factory_code} 不存在")
            if run.factory_id != factory.id:
                raise ValueError(f"生产事件 {run_no} 的工厂与导入内容不一致")
        if vehicle_model_code:
            vehicle_model = vehicle_models_by_code.get(vehicle_model_code)
            if not vehicle_model:
                raise ValueError(f"车型 {vehicle_model_code} 不存在")
            if run.vehicle_model_id != vehicle_model.id:
                raise ValueError(f"生产事件 {run_no} 的车型与导入内容不一致")
        if color_code:
            color = colors_by_code.get(color_code)
            if not color:
                raise ValueError(f"颜色 {color_code} 不存在")
            if run.color_id != color.id:
                raise ValueError(f"生产事件 {run_no} 的颜色与导入内容不一致")
        return run

    if not body_no or not factory_code or not vehicle_model_code or not color_code:
        raise ValueError(
            "生产事件不存在时，必须同时填写车号、工厂代码、车型代码和颜色代码"
        )

    factory = factories_by_code.get(factory_code)
    if not factory:
        raise ValueError(f"工厂 {factory_code} 不存在")
    vehicle_model = vehicle_models_by_code.get(vehicle_model_code)
    if not vehicle_model:
        raise ValueError(f"车型 {vehicle_model_code} 不存在")
    color = colors_by_code.get(color_code)
    if not color:
        raise ValueError(f"颜色 {color_code} 不存在")

    run = create_production_run(
        ProductionRunCreate(
            run_no=run_no,
            body_no=body_no,
            factory_id=factory.id,
            vehicle_model_id=vehicle_model.id,
            color_id=color.id,
            shift=shift,
            started_at=_coerce_datetime(started_at_value)
            if not _is_blank(started_at_value)
            else _coerce_datetime(measured_at_value),
            completed_at=_coerce_datetime(completed_at_value)
            if not _is_blank(completed_at_value)
            else None,
        ),
        db,
    )
    runs_by_no[run_no] = run
    return run


def _quality_measurement_template_rows(
    db: Session | None,
    *,
    quality_type: str | None = None,
    factory_code: str | None = None,
    color_code: str | None = None,
    vehicle_model_code: str | None = None,
    shift: str | None = None,
) -> list[dict[str, Any]]:
    if db is None:
        return []
    query = (
        select(MeasurementGroupPoint, MeasurementGroup, MeasurementPoint, VehicleModel, Part)
        .join(MeasurementGroup, MeasurementGroup.id == MeasurementGroupPoint.measurement_group_id)
        .join(MeasurementPoint, MeasurementPoint.id == MeasurementGroupPoint.measurement_point_id)
        .join(VehicleModel, VehicleModel.id == MeasurementGroup.vehicle_model_id)
        .join(Part, Part.id == MeasurementPoint.part_id)
        .order_by(MeasurementGroup.code, MeasurementGroupPoint.sequence_no, MeasurementPoint.code)
    )
    normalized = quality_type.strip().upper() if quality_type else None
    if normalized:
        _quality_metric_catalog_for(normalized)
        query = query.where(MeasurementGroup.quality_type == normalized)
    preferred_model = _normalize_optional_text(vehicle_model_code)
    if preferred_model:
        query = query.where(VehicleModel.code == preferred_model)
    pref_factory = _normalize_optional_text(factory_code) or ""
    pref_color = _normalize_optional_text(color_code) or ""
    pref_shift = _normalize_optional_text(shift) or ""
    rows: list[dict[str, Any]] = []
    for _, group, point, model, part in db.execute(query):
        rows.append(
            {
                "production_run_no": "",
                "body_no": "",
                "factory_code": pref_factory,
                "measurement_group_code": group.code,
                "measurement_group_name": group.name,
                "measurement_point_code": point.code,
                "measurement_point_name": point.name,
                "vehicle_model_code": model.code,
                "vehicle_model_name": model.name,
                "part_code": part.code,
                "part_name": part.name,
                "quality_type": group.quality_type,
                "color_code": pref_color,
                "shift": pref_shift,
                "production_started_at": "",
                "production_completed_at": "",
                "measured_at": "",
                "measured_by": "",
                "data_type": "TEST",
                "device_code": "",
                "instrument_code": "",
                "measurement_method_code": "",
                "calibration_no": "",
                "reference_standard_code": "",
                "import_profile_code": "",
                "measurement_direction": "",
                "raw_file_uri": "",
                "status_score": "",
                "is_valid": True,
            }
        )
    return rows


def _quality_measurement_export_rows(
    db: Session,
    *,
    quality_type: str | None = None,
) -> list[dict[str, Any]]:
    normalized = quality_type.strip().upper() if quality_type else None
    if normalized:
        _quality_metric_catalog_for(normalized)
    query = select(QualityMeasurement).order_by(QualityMeasurement.measured_at)
    if normalized:
        query = query.where(QualityMeasurement.quality_type == normalized)
    measurements = list(db.scalars(query))
    if not measurements:
        return []

    run_map = {
        item.id: item
        for item in db.scalars(
            select(ProductionRun).where(
                ProductionRun.id.in_({measurement.production_run_id for measurement in measurements})
            )
        )
    }
    point_map = {
        item.id: item
        for item in db.scalars(
            select(MeasurementPoint).where(
                MeasurementPoint.id.in_({measurement.measurement_point_id for measurement in measurements})
            )
        )
    }
    group_ids = {
        measurement.measurement_group_id
        for measurement in measurements
        if measurement.measurement_group_id
    }
    group_map = {
        item.id: item
        for item in db.scalars(select(MeasurementGroup).where(MeasurementGroup.id.in_(group_ids)))
    } if group_ids else {}
    model_ids = {
        point.vehicle_model_id
        for point in point_map.values()
        if point.vehicle_model_id
    }
    model_map = {
        item.id: item
        for item in db.scalars(select(VehicleModel).where(VehicleModel.id.in_(model_ids)))
    } if model_ids else {}
    part_ids = {point.part_id for point in point_map.values() if point.part_id}
    part_map = {
        item.id: item
        for item in db.scalars(select(Part).where(Part.id.in_(part_ids)))
    } if part_ids else {}
    instrument_ids = {
        measurement.instrument_id for measurement in measurements if measurement.instrument_id
    }
    instrument_map = {
        item.id: item
        for item in db.scalars(
            select(MeasurementInstrument).where(MeasurementInstrument.id.in_(instrument_ids))
        )
    } if instrument_ids else {}
    method_ids = {
        measurement.measurement_method_id
        for measurement in measurements
        if measurement.measurement_method_id
    }
    method_map = {
        item.id: item
        for item in db.scalars(select(MeasurementMethod).where(MeasurementMethod.id.in_(method_ids)))
    } if method_ids else {}
    calibration_ids = {
        measurement.calibration_record_id
        for measurement in measurements
        if measurement.calibration_record_id
    }
    calibration_map = {
        item.id: item
        for item in db.scalars(
            select(MeasurementCalibrationRecord).where(
                MeasurementCalibrationRecord.id.in_(calibration_ids)
            )
        )
    } if calibration_ids else {}
    reference_ids = {
        measurement.reference_standard_id
        for measurement in measurements
        if measurement.reference_standard_id
    }
    reference_map = {
        item.id: item
        for item in db.scalars(
            select(MeasurementReferenceStandard).where(
                MeasurementReferenceStandard.id.in_(reference_ids)
            )
        )
    } if reference_ids else {}
    profile_ids = {
        measurement.import_profile_id
        for measurement in measurements
        if measurement.import_profile_id
    }
    profile_map = {
        item.id: item
        for item in db.scalars(
            select(MeasurementImportProfile).where(MeasurementImportProfile.id.in_(profile_ids))
        )
    } if profile_ids else {}
    factory_ids = {
        run.factory_id
        for run in run_map.values()
        if run.factory_id
    }
    factory_map = {
        item.id: item
        for item in db.scalars(select(Factory).where(Factory.id.in_(factory_ids)))
    } if factory_ids else {}
    color_ids = {
        run.color_id
        for run in run_map.values()
        if run.color_id
    }
    color_map = {
        item.id: item
        for item in db.scalars(select(Color).where(Color.id.in_(color_ids)))
    } if color_ids else {}
    metric_rows: dict[str, dict[str, float]] = {}
    metric_query = select(QualityMetricValue).where(
        QualityMetricValue.measurement_id.in_({measurement.id for measurement in measurements})
    )
    for metric in db.scalars(metric_query):
        metric_rows.setdefault(metric.measurement_id, {})[metric.metric_code] = metric.raw_value

    rows: list[dict[str, Any]] = []
    for measurement in measurements:
        point = point_map.get(measurement.measurement_point_id)
        group = group_map.get(measurement.measurement_group_id) if measurement.measurement_group_id else None
        model = model_map.get(point.vehicle_model_id) if point else None
        part = part_map.get(point.part_id) if point else None
        run = run_map.get(measurement.production_run_id)
        factory = factory_map.get(run.factory_id) if run else None
        color = color_map.get(run.color_id) if run else None
        instrument = instrument_map.get(measurement.instrument_id) if measurement.instrument_id else None
        method = method_map.get(measurement.measurement_method_id) if measurement.measurement_method_id else None
        calibration = calibration_map.get(measurement.calibration_record_id) if measurement.calibration_record_id else None
        reference = reference_map.get(measurement.reference_standard_id) if measurement.reference_standard_id else None
        profile = profile_map.get(measurement.import_profile_id) if measurement.import_profile_id else None
        row = {
            "id": measurement.id,
            "data_no": measurement.data_no,
            "production_run_no": run.run_no if run else "",
            "body_no": run.body_no if run else "",
            "factory_code": factory.code if factory else "",
            "measurement_group_code": group.code if group else "",
            "measurement_group_name": group.name if group else "",
            "measurement_point_code": point.code if point else "",
            "measurement_point_name": point.name if point else "",
            "vehicle_model_code": model.code if model else "",
            "vehicle_model_name": model.name if model else "",
            "part_code": part.code if part else "",
            "part_name": part.name if part else "",
            "quality_type": measurement.quality_type,
            "color_code": color.code if color else "",
            "shift": run.shift if run else "",
            "production_started_at": run.started_at if run else "",
            "production_completed_at": run.completed_at if run else "",
            "measured_at": measurement.measured_at,
            "measured_by": measurement.measured_by,
            "data_type": measurement.data_type,
            "device_code": measurement.device_code,
            "instrument_code": instrument.code if instrument else "",
            "measurement_method_code": f"{method.code}:{method.version}" if method else "",
            "calibration_no": calibration.calibration_no if calibration else "",
            "reference_standard_code": reference.code if reference else "",
            "import_profile_code": f"{profile.code}:{profile.version}" if profile else "",
            "measurement_direction": measurement.measurement_direction,
            "raw_file_uri": measurement.raw_file_uri,
            "status_score": measurement.status_score,
            "is_valid": measurement.is_valid,
        }
        for metric in _quality_metric_catalog_for(normalized or measurement.quality_type):
            metric_value = metric_rows.get(measurement.id, {}).get(metric["code"])
            if metric_value is not None:
                row[_quality_metric_field_name(metric["code"])] = metric_value
        rows.append(row)
    return rows


def _resolve_lookup_by_code(
    value: Any,
    mapping: dict[str, Any],
    *,
    label: str,
) -> str | None:
    if _is_blank(value):
        return None
    code = str(value).strip()
    item = mapping.get(code)
    if not item:
        raise ValueError(f"{label} {code} 不存在")
    return item.id


def _resolve_lookup_by_code_or_version(
    value: Any,
    mapping: dict[str, list[Any]],
    *,
    label: str,
) -> str | None:
    if _is_blank(value):
        return None
    text = str(value).strip()
    if ":" in text:
        code, version = [segment.strip() for segment in text.split(":", 1)]
        matches = [item for item in mapping.get(code, []) if getattr(item, "version", "") == version]
        if not matches:
            raise ValueError(f"{label} {text} 不存在")
        return matches[0].id
    matches = mapping.get(text, [])
    if not matches:
        raise ValueError(f"{label} {text} 不存在")
    if len(matches) > 1:
        raise ValueError(f"{label} {text} 存在多个版本，请填写 code:version")
    return matches[0].id


def _optional_number(value: Any) -> float | None:
    if _is_blank(value):
        return None
    return float(value)


def _resolve_quality_measurement_existing_id(db: Session, row: dict[str, Any], data_no: str) -> str | None:
    row_id = row.get("id")
    if row_id and db.get(QualityMeasurement, row_id):
        return str(row_id)
    existing = db.scalar(select(QualityMeasurement).where(QualityMeasurement.data_no == data_no))
    return existing.id if existing else None


def _import_quality_measurements(
    content: bytes,
    *,
    filename: str,
    mode: ImportMode,
    progress_callback: Callable[[], None] | None,
    db: Session,
) -> dict[str, Any]:
    quality_fields = _quality_measurement_bulk_fields()
    rows = _parse_rows(content, filename, fields=quality_fields)
    allowed_fields = {field.name for field in quality_fields}
    unknown = sorted({key for row in rows for key in row if key and key not in allowed_fields})
    if unknown:
        raise HTTPException(status_code=422, detail=f"模板字段不匹配，未知字段：{', '.join(unknown)}")

    runs_by_no = {item.run_no: item for item in db.scalars(select(ProductionRun))}
    factories_by_code = {item.code: item for item in db.scalars(select(Factory))}
    vehicle_models_by_code = {item.code: item for item in db.scalars(select(VehicleModel))}
    colors_by_code = {item.code: item for item in db.scalars(select(Color))}
    groups_by_vehicle_code = {
        (item.vehicle_model_id, item.code): item for item in db.scalars(select(MeasurementGroup))
    }
    points_by_vehicle_code = {
        (item.vehicle_model_id, item.code): item for item in db.scalars(select(MeasurementPoint))
    }
    instruments_by_code = {
        item.code: item for item in db.scalars(select(MeasurementInstrument))
    }
    methods_by_code: dict[str, list[Any]] = {}
    for item in db.scalars(select(MeasurementMethod)):
        methods_by_code.setdefault(item.code, []).append(item)
    calibrations_by_no = {
        item.calibration_no: item for item in db.scalars(select(MeasurementCalibrationRecord))
    }
    references_by_code = {
        item.code: item for item in db.scalars(select(MeasurementReferenceStandard))
    }
    profiles_by_code: dict[str, list[Any]] = {}
    for item in db.scalars(select(MeasurementImportProfile)):
        profiles_by_code.setdefault(item.code, []).append(item)

    created = 0
    updated = 0
    skipped = 0
    errors: list[dict[str, Any]] = []
    for row_index, raw_row in enumerate(rows, start=2):
        if progress_callback:
            progress_callback()
        if not any(str(value or "").strip() for value in raw_row.values()):
            skipped += 1
            continue
        try:
            data_no = str(raw_row.get("data_no") or "").strip()
            run_no = str(raw_row.get("production_run_no") or "").strip()
            body_no = str(raw_row.get("body_no") or "").strip()
            point_code = str(raw_row.get("measurement_point_code") or "").strip()
            measured_at_value = raw_row.get("measured_at")
            if not point_code or _is_blank(measured_at_value):
                raise ValueError("请填写测量点代码和测量时间")
            if not run_no and not body_no:
                raise ValueError("请填写生产记录编号，或填写车号让系统自动生成生产记录")

            run = _resolve_or_create_quality_production_run(
                db=db,
                runs_by_no=runs_by_no,
                factories_by_code=factories_by_code,
                vehicle_models_by_code=vehicle_models_by_code,
                colors_by_code=colors_by_code,
                row=raw_row,
                measured_at_value=measured_at_value,
            )

            group_code = str(raw_row.get("measurement_group_code") or "").strip()
            group = (
                groups_by_vehicle_code.get((run.vehicle_model_id, group_code))
                if group_code
                else None
            )
            if group_code and not group:
                raise ValueError(f"测量编组 {group_code} 不存在或不属于当前生产事件车型")

            point = points_by_vehicle_code.get((run.vehicle_model_id, point_code))
            if not point:
                raise ValueError(f"测量点 {point_code} 不存在或不属于当前生产事件车型")

            quality_type = str(
                _translate_business_value("quality_type", raw_row.get("quality_type") or "")
            ).strip().upper()
            if not quality_type and group:
                quality_type = str(group.quality_type).strip().upper()
            if not quality_type:
                if len(point.quality_types or []) == 1:
                    quality_type = str(point.quality_types[0]).upper()
                else:
                    raise ValueError("请选择质量指标类型")
            metric_catalog = {item["code"]: item for item in _quality_metric_catalog_for(quality_type)}
            metrics = []
            for name, raw_value in raw_row.items():
                metric_code = _quality_metric_code_from_field(name)
                if not metric_code or _is_blank(raw_value):
                    continue
                definition = metric_catalog.get(metric_code)
                if not definition:
                    raise ValueError(f"质量类型 {quality_type} 不支持指标 {metric_code}")
                metrics.append(
                    {
                        "metric_code": metric_code,
                        "metric_name": definition["name"],
                        "raw_value": float(raw_value),
                        "unit": definition.get("unit"),
                    }
                )
            if not metrics:
                raise ValueError("至少填写 1 个质量指标列")

            if not data_no:
                resolved_body = body_no or str(run.body_no or "").strip()
                if not resolved_body:
                    raise ValueError("自动生成质量数据编号需要车号；请填写车号，或选择已维护车号的生产记录")
                data_no = _build_quality_data_no(
                    body_no=resolved_body,
                    point_code=point_code,
                    quality_type=quality_type,
                )

            payload = {
                "data_no": data_no,
                "production_run_id": run.id,
                "measurement_group_id": group.id if group else None,
                "measurement_point_id": point.id,
                "quality_type": quality_type,
                "data_type": str(
                    _translate_business_value("data_type", raw_row.get("data_type") or "TEST")
                ).strip() or "TEST",
                "measured_at": _coerce_datetime(measured_at_value),
                "measured_by": str(raw_row.get("measured_by") or "").strip() or None,
                "device_code": str(raw_row.get("device_code") or "").strip() or None,
                "instrument_id": _resolve_lookup_by_code(
                    raw_row.get("instrument_code"),
                    instruments_by_code,
                    label="仪器",
                ),
                "measurement_method_id": _resolve_lookup_by_code_or_version(
                    raw_row.get("measurement_method_code"),
                    methods_by_code,
                    label="测量方法",
                ),
                "calibration_record_id": _resolve_lookup_by_code(
                    raw_row.get("calibration_no"),
                    calibrations_by_no,
                    label="校准记录",
                ),
                "reference_standard_id": _resolve_lookup_by_code(
                    raw_row.get("reference_standard_code"),
                    references_by_code,
                    label="参考件",
                ),
                "import_profile_id": _resolve_lookup_by_code_or_version(
                    raw_row.get("import_profile_code"),
                    profiles_by_code,
                    label="导入模板",
                ),
                "measurement_direction": str(raw_row.get("measurement_direction") or "").strip() or None,
                "raw_file_uri": str(raw_row.get("raw_file_uri") or "").strip() or None,
                "status_score": _optional_number(raw_row.get("status_score")),
                "is_valid": True if _is_blank(raw_row.get("is_valid")) else _coerce_bool(raw_row.get("is_valid")),
                "metrics": metrics,
                "repeat_readings": [],
            }

            existing_id = _resolve_quality_measurement_existing_id(db, raw_row, data_no) if mode == "upsert" else None
            if existing_id:
                update_quality_measurement(existing_id, QualityMeasurementUpdate(**payload), db)
                updated += 1
            else:
                create_quality_measurement(QualityMeasurementCreate(**payload), db)
                created += 1
        except Exception as exc:  # noqa: BLE001 - row-level import must collect all failures.
            db.rollback()
            errors.append({"row": row_index, "message": _error_message(exc)})

    return {
        "resource_key": "quality.measurements",
        "resource_label": "质量测量",
        "mode": mode,
        "total_rows": len(rows),
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "failed": len(errors),
        "errors": errors[:100],
        "truncated_errors": len(errors) > 100,
    }


def _brush_contribution_bulk_fields() -> tuple[BulkField, ...]:
    """Accepted brush contribution columns, including legacy internal references."""
    return (
        BulkField(
            "id",
            "贡献记录 ID",
            "string",
            False,
            "",
            "可选；导出回写时使用，新建可留空",
        ),
        BulkField(
            "factory_code",
            "工厂代码",
            "string",
            False,
            "F01",
            "只读上下文，导入时忽略；由当前刷子所属程序自动填写",
        ),
        BulkField(
            "factory_name",
            "工厂名称",
            "string",
            False,
            "",
            "只读上下文，导入时忽略",
        ),
        BulkField(
            "program_code",
            "喷涂程序代码",
            "string",
            False,
            "PRG-CC2",
            "与 program_version + brush_no 一起定位刷子；也可直接填 brush_id",
        ),
        BulkField(
            "program_name",
            "喷涂程序名称",
            "string",
            False,
            "",
            "只读上下文，导入时忽略",
        ),
        BulkField(
            "program_version",
            "程序版本号",
            "string",
            False,
            "V1",
            "与 program_code + brush_no 一起定位刷子",
        ),
        BulkField(
            "process_stage",
            "工序阶段",
            "string",
            False,
            "CLEARCOAT_2",
            "只读上下文，导入时忽略",
        ),
        BulkField(
            "brush_no",
            "刷子号",
            "string",
            False,
            "B01",
            "与程序版本一起定位刷子；页面下载模板时已预填",
        ),
        BulkField(
            "brush_table_no",
            "刷子表号",
            "string",
            False,
            "BT01",
            "只读上下文，导入时忽略",
        ),
        BulkField(
            "brush_id",
            "刷子 ID",
            "string",
            False,
            "",
            "可选；留空时按 factory/program/version/brush_no 解析，或使用导入默认 brush_id",
        ),
        BulkField(
            "vehicle_model_code",
            "车型代码",
            "string",
            True,
            "VM01",
            "与 measurement_point_code 一起定位测量点",
        ),
        BulkField(
            "vehicle_model_name",
            "车型名称",
            "string",
            False,
            "",
            "只读上下文，导入时忽略",
        ),
        BulkField(
            "part_code",
            "零件代码",
            "string",
            False,
            "ROOF",
            "只读上下文，导入时忽略",
        ),
        BulkField(
            "part_name",
            "零件名称",
            "string",
            False,
            "",
            "只读上下文，导入时忽略",
        ),
        BulkField(
            "measurement_point_code",
            "测量点代码",
            "string",
            True,
            "P01",
            "与车型代码一起定位质量测量点",
        ),
        BulkField(
            "measurement_point_name",
            "测量点名称",
            "string",
            False,
            "",
            "只读上下文，导入时忽略",
        ),
        BulkField(
            "measurement_point_id",
            "测量点 ID",
            "string",
            False,
            "",
            "可选；留空时按车型+测量点代码解析",
        ),
        BulkField(
            "overlap_ratio",
            "重叠率",
            "number",
            True,
            "0.5",
            "0~1，填写本刷子对该点的重叠率",
        ),
        BulkField(
            "contribution_weight",
            "贡献权重",
            "number",
            True,
            "0.5",
            "0~1（不含 0），填写本刷子对该点的贡献权重",
        ),
        BulkField("source", "来源", "string", False, "EXPERT", "默认 EXPERT"),
        BulkField("version", "贡献版本", "string", False, "1.0", "默认 1.0"),
        BulkField("is_approved", "是否已审批", "boolean", False, "false", "true/false"),
    )


def _brush_contribution_template_fields() -> tuple[BulkField, ...]:
    """Business-facing columns; parent brush and point are resolved from context/codes."""
    hidden = {"id", "brush_id", "measurement_point_id"}
    return tuple(
        field for field in _brush_contribution_bulk_fields() if field.name not in hidden
    )


def _brush_contribution_template_rows(
    db: Session | None,
    *,
    brush_id: str | None = None,
) -> list[dict[str, Any]]:
    if db is None or not brush_id:
        return []
    brush = db.get(Brush, brush_id)
    if not brush:
        raise HTTPException(status_code=404, detail="刷子不存在")
    version = db.get(SprayProgramVersion, brush.program_version_id)
    if not version:
        raise HTTPException(status_code=404, detail="程序版本不存在")
    program = db.get(SprayProgram, version.spray_program_id)
    if not program:
        raise HTTPException(status_code=404, detail="喷涂程序不存在")
    factory = db.get(Factory, program.factory_id)
    if not factory:
        raise HTTPException(status_code=404, detail="工厂不存在")

    applicable_vehicle_model_ids = set(
        db.scalars(
            select(ProgramVehicleModel.vehicle_model_id).where(
                ProgramVehicleModel.program_version_id == version.id
            )
        )
    )
    existing_by_point = {
        item.measurement_point_id: item
        for item in db.scalars(
            select(BrushPointContribution).where(BrushPointContribution.brush_id == brush.id)
        )
    }

    query = (
        select(MeasurementPoint, VehicleModel, Part)
        .join(VehicleModel, VehicleModel.id == MeasurementPoint.vehicle_model_id)
        .join(Part, Part.id == MeasurementPoint.part_id)
        .where(MeasurementPoint.point_type == "QUALITY")
        .order_by(VehicleModel.code, MeasurementPoint.code)
    )
    if brush.part_id:
        query = query.where(MeasurementPoint.part_id == brush.part_id)
    if applicable_vehicle_model_ids:
        query = query.where(MeasurementPoint.vehicle_model_id.in_(applicable_vehicle_model_ids))

    rows: list[dict[str, Any]] = []
    for point, model, part in db.execute(query):
        existing = existing_by_point.get(point.id)
        rows.append(
            {
                "factory_code": factory.code,
                "factory_name": factory.name,
                "program_code": program.program_code,
                "program_name": program.name,
                "program_version": version.version,
                "process_stage": program.process_stage,
                "brush_no": brush.brush_no,
                "brush_table_no": brush.brush_table_no,
                "brush_id": brush.id,
                "vehicle_model_code": model.code,
                "vehicle_model_name": model.name,
                "part_code": part.code,
                "part_name": part.name,
                "measurement_point_code": point.code,
                "measurement_point_name": point.name,
                "measurement_point_id": point.id,
                "overlap_ratio": existing.overlap_ratio if existing else "",
                "contribution_weight": existing.contribution_weight if existing else "",
                "source": existing.source if existing else "EXPERT",
                "version": existing.version if existing else "1.0",
                "is_approved": existing.is_approved if existing else False,
            }
        )
    return rows


def _resolve_brush_for_contribution_import(
    row: dict[str, Any],
    *,
    brushes_by_id: dict[str, Brush],
    programs_by_factory_code: dict[tuple[str, str], SprayProgram],
    versions_by_program_version: dict[tuple[str, str], SprayProgramVersion],
    brushes_by_version_no: dict[tuple[str, str], Brush],
    factories_by_code: dict[str, Factory],
) -> Brush:
    brush_id = _normalize_optional_text(row.get("brush_id"))
    if brush_id:
        brush = brushes_by_id.get(brush_id)
        if not brush:
            raise ValueError(f"刷子 {brush_id} 不存在")
        return brush

    program_code = _normalize_optional_text(row.get("program_code"))
    program_version = _normalize_optional_text(row.get("program_version"))
    brush_no = _normalize_optional_text(row.get("brush_no"))
    if not program_code or not program_version or not brush_no:
        raise ValueError(
            "缺少刷子定位信息：请填写 brush_id，或同时填写 program_code、program_version、brush_no"
        )

    factory_code = _normalize_optional_text(row.get("factory_code"))
    if factory_code:
        factory = factories_by_code.get(factory_code)
        if not factory:
            raise ValueError(f"工厂 {factory_code} 不存在")
        program = programs_by_factory_code.get((factory.id, program_code))
        if not program:
            raise ValueError(f"工厂 {factory_code} 下喷涂程序 {program_code} 不存在")
    else:
        matches = [
            program
            for (_fid, code), program in programs_by_factory_code.items()
            if code == program_code
        ]
        if not matches:
            raise ValueError(f"喷涂程序 {program_code} 不存在")
        if len(matches) > 1:
            raise ValueError(
                f"喷涂程序代码 {program_code} 在多个工厂下存在，请同时填写 factory_code"
            )
        program = matches[0]

    version = versions_by_program_version.get((program.id, program_version))
    if not version:
        raise ValueError(f"程序 {program_code} 的版本 {program_version} 不存在")
    brush = brushes_by_version_no.get((version.id, brush_no))
    if not brush:
        raise ValueError(f"程序版本 {program_code}/{program_version} 下刷子 {brush_no} 不存在")
    return brush


def _resolve_measurement_point_for_contribution_import(
    row: dict[str, Any],
    *,
    points_by_id: dict[str, MeasurementPoint],
    points_by_vehicle_code: dict[tuple[str, str], MeasurementPoint],
    vehicle_models_by_code: dict[str, VehicleModel],
) -> MeasurementPoint:
    point_id = _normalize_optional_text(row.get("measurement_point_id"))
    if point_id:
        point = points_by_id.get(point_id)
        if not point:
            raise ValueError(f"测量点 {point_id} 不存在")
        return point

    point_code = _normalize_optional_text(row.get("measurement_point_code"))
    vehicle_model_code = _normalize_optional_text(row.get("vehicle_model_code"))
    if not point_code:
        raise ValueError("缺少必填字段：measurement_point_code 或 measurement_point_id")
    if not vehicle_model_code:
        raise ValueError("按测量点代码导入时必须填写 vehicle_model_code")
    vehicle_model = vehicle_models_by_code.get(vehicle_model_code)
    if not vehicle_model:
        raise ValueError(f"车型 {vehicle_model_code} 不存在")
    point = points_by_vehicle_code.get((vehicle_model.id, point_code))
    if not point:
        raise ValueError(f"车型 {vehicle_model_code} 下测量点 {point_code} 不存在")
    return point


def _import_brush_contributions(
    content: bytes,
    *,
    filename: str,
    mode: ImportMode,
    default_values: dict[str, Any] | None,
    progress_callback: Callable[[], None] | None,
    db: Session,
) -> dict[str, Any]:
    resource = get_resource("process.brush-contributions")
    contribution_fields = _brush_contribution_bulk_fields()
    rows = _parse_rows(content, filename, fields=contribution_fields)
    allowed_fields = {field.name for field in contribution_fields}
    unknown = sorted({key for row in rows for key in row if key and key not in allowed_fields})
    if unknown:
        raise HTTPException(status_code=422, detail=f"模板字段不匹配，未知字段：{', '.join(unknown)}")

    default_values = default_values or {}
    allowed_defaults = {"brush_id"}
    unknown_defaults = sorted(key for key in default_values if key not in allowed_defaults)
    if unknown_defaults:
        raise HTTPException(
            status_code=422,
            detail=f"默认导入字段不匹配：{', '.join(unknown_defaults)}",
        )

    factories_by_code = {item.code: item for item in db.scalars(select(Factory))}
    programs = list(db.scalars(select(SprayProgram)))
    programs_by_factory_code = {(item.factory_id, item.program_code): item for item in programs}
    versions = list(db.scalars(select(SprayProgramVersion)))
    versions_by_program_version = {
        (item.spray_program_id, item.version): item for item in versions
    }
    brushes = list(db.scalars(select(Brush)))
    brushes_by_id = {item.id: item for item in brushes}
    brushes_by_version_no = {(item.program_version_id, item.brush_no): item for item in brushes}
    vehicle_models_by_code = {item.code: item for item in db.scalars(select(VehicleModel))}
    points = list(db.scalars(select(MeasurementPoint)))
    points_by_id = {item.id: item for item in points}
    points_by_vehicle_code = {(item.vehicle_model_id, item.code): item for item in points}

    created = 0
    updated = 0
    skipped = 0
    errors: list[dict[str, Any]] = []
    data_keys = {
        "brush_id",
        "measurement_point_id",
        "measurement_point_code",
        "vehicle_model_code",
        "overlap_ratio",
        "contribution_weight",
        "source",
        "version",
        "is_approved",
        "program_code",
        "program_version",
        "brush_no",
    }
    for row_index, raw_row in enumerate(rows, start=2):
        if progress_callback:
            progress_callback()
        merged_row = _apply_default_values(raw_row, default_values)
        if not any(str(merged_row.get(key) or "").strip() for key in data_keys):
            skipped += 1
            continue
        try:
            brush = _resolve_brush_for_contribution_import(
                merged_row,
                brushes_by_id=brushes_by_id,
                programs_by_factory_code=programs_by_factory_code,
                versions_by_program_version=versions_by_program_version,
                brushes_by_version_no=brushes_by_version_no,
                factories_by_code=factories_by_code,
            )
            point = _resolve_measurement_point_for_contribution_import(
                merged_row,
                points_by_id=points_by_id,
                points_by_vehicle_code=points_by_vehicle_code,
                vehicle_models_by_code=vehicle_models_by_code,
            )
            if _is_blank(merged_row.get("overlap_ratio")) or _is_blank(
                merged_row.get("contribution_weight")
            ):
                raise ValueError("请填写重叠率和贡献权重")

            payload = BrushPointContributionUpsert(
                overlap_ratio=float(merged_row["overlap_ratio"]),
                contribution_weight=float(merged_row["contribution_weight"]),
                source=str(
                    _translate_business_value("source", merged_row.get("source") or "EXPERT")
                ).strip() or "EXPERT",
                version=str(merged_row.get("version") or "1.0").strip() or "1.0",
                is_approved=(
                    False
                    if _is_blank(merged_row.get("is_approved"))
                    else _coerce_bool(merged_row.get("is_approved"))
                ),
            )
            existing_id = (
                _brush_contribution_match(
                    db,
                    {"brush_id": brush.id, "measurement_point_id": point.id},
                )
                if mode == "upsert"
                else None
            )
            upsert_brush_point_contribution(brush.id, point.id, payload, db)
            if existing_id:
                updated += 1
            else:
                created += 1
        except Exception as exc:  # noqa: BLE001 - row-level import must collect all failures.
            db.rollback()
            errors.append({"row": row_index, "message": _error_message(exc)})

    return {
        "resource_key": resource.key,
        "resource_label": resource.label,
        "mode": mode,
        "total_rows": len(rows),
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "failed": len(errors),
        "errors": errors[:100],
        "truncated_errors": len(errors) > 100,
    }


def _brush_contribution_match(db: Session, row: dict[str, Any]) -> str | None:
    if _is_blank(row.get("brush_id")) or _is_blank(row.get("measurement_point_id")):
        return None
    existing = db.scalar(
        select(BrushPointContribution).where(
            BrushPointContribution.brush_id == row["brush_id"],
            BrushPointContribution.measurement_point_id == row["measurement_point_id"],
        )
    )
    return existing.id if existing else None


def _contribution_entry_match(db: Session, row: dict[str, Any]) -> str | None:
    if _is_blank(row.get("contribution_version_id")) or _is_blank(row.get("measurement_point_id")):
        return None
    source_key = None
    if row.get("brush_id"):
        source_key = f"BRUSH:{row['brush_id']}"
    if row.get("path_segment_id"):
        source_key = f"PATH:{row['path_segment_id']}"
    if not source_key:
        return None
    existing = db.scalar(
        select(PointContributionEntry).where(
            PointContributionEntry.contribution_version_id == row["contribution_version_id"],
            PointContributionEntry.measurement_point_id == row["measurement_point_id"],
            PointContributionEntry.source_key == source_key,
        )
    )
    return existing.id if existing else None


def _bind_program_vehicle_model(
    payload: ProgramVehicleModelBulkCreate,
    db: Session,
) -> ProgramVehicleModel:
    check_fk(db, SprayProgramVersion, payload.program_version_id, label="喷涂程序版本")
    check_fk(db, VehicleModel, payload.vehicle_model_id, label="车型")
    existing = db.scalar(
        select(ProgramVehicleModel).where(
            ProgramVehicleModel.program_version_id == payload.program_version_id,
            ProgramVehicleModel.vehicle_model_id == payload.vehicle_model_id,
        )
    )
    if existing:
        raise HTTPException(status_code=409, detail="程序版本已关联该车型")
    resource = ProgramVehicleModel(**payload.model_dump())
    db.add(resource)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="程序版本已关联该车型") from exc
    db.refresh(resource)
    return resource


def _bind_program_color(payload: ProgramColorBulkCreate, db: Session) -> ProgramColor:
    check_fk(db, SprayProgramVersion, payload.program_version_id, label="喷涂程序版本")
    check_fk(db, Color, payload.color_id, label="颜色")
    existing = db.scalar(
        select(ProgramColor).where(
            ProgramColor.program_version_id == payload.program_version_id,
            ProgramColor.color_id == payload.color_id,
        )
    )
    if existing:
        raise HTTPException(status_code=409, detail="程序版本已关联该颜色")
    resource = ProgramColor(**payload.model_dump())
    db.add(resource)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="程序版本已关联该颜色") from exc
    db.refresh(resource)
    return resource


def _body_map_3d_export_row(
    db: Session,
    layout: MeasurementPoint3DLayout,
) -> dict[str, Any]:
    del db
    return {
        "id": layout.id,
        "measurement_point_id": layout.measurement_point_id,
        "pos_x": layout.pos_x,
        "pos_y": layout.pos_y,
        "pos_z": layout.pos_z,
        "normal_x": layout.normal_x,
        "normal_y": layout.normal_y,
        "normal_z": layout.normal_z,
        "model_asset_key": layout.model_asset_key,
        "project_to_2d": False,
    }


def _resource(
    key: str,
    label: str,
    group: str,
    model: type,
    create_schema: type[BaseModel],
    update_schema: type[BaseModel] | None,
    create_func_ref: Callable[..., Any],
    update_func_ref: Callable[..., Any] | None = None,
    *,
    unique_fields: tuple[str, ...] = (),
    create_arg_fields: tuple[str, ...] = (),
    update_arg_fields: tuple[str, ...] = (),
    update_uses_resource_id: bool = True,
    order_by: tuple[str, ...] = ("created_at",),
    extra_fields: tuple[BulkField, ...] = (),
    export_row: Callable[[Session, Any], dict[str, Any]] | None = None,
    match_existing: Callable[[Session, dict[str, Any]], str | None] | None = None,
    importable: bool = True,
) -> BulkResource:
    return BulkResource(
        key=key,
        label=label,
        group=group,
        model=model,
        create_schema=create_schema,
        update_schema=update_schema,
        create_func=create_func_ref,
        update_func=update_func_ref,
        unique_fields=unique_fields,
        create_arg_fields=create_arg_fields,
        update_arg_fields=update_arg_fields,
        update_uses_resource_id=update_uses_resource_id,
        order_by=order_by,
        extra_fields=extra_fields,
        export_row=export_row,
        match_existing=match_existing,
        importable=importable,
    )


RESOURCES: dict[str, BulkResource] = {
    resource.key: resource
    for resource in [
        _resource("master.factories", "工厂", "主数据", Factory, FactoryCreate, FactoryUpdate, create_factory, update_factory, unique_fields=("code",), order_by=("code",)),
        _resource("master.vehicle-models", "车型", "主数据", VehicleModel, VehicleModelCreate, VehicleModelUpdate, create_vehicle_model, update_vehicle_model, unique_fields=("code",), order_by=("code",)),
        _resource("master.colors", "颜色", "主数据", Color, ColorCreate, ColorUpdate, create_color, update_color, unique_fields=("code",), order_by=("code",)),
        _resource("master.parts", "零件", "主数据", Part, PartCreate, PartUpdate, create_part, update_part, unique_fields=("code",), order_by=("code",)),
        _resource("master.measurement-groups", "测量编组", "主数据", MeasurementGroup, MeasurementGroupCreate, MeasurementGroupUpdate, create_measurement_group, update_measurement_group, unique_fields=("vehicle_model_id", "code"), order_by=("code",)),
        _resource("master.measurement-points", "测量点", "主数据", MeasurementPoint, MeasurementPointCreate, MeasurementPointUpdate, create_measurement_point, update_measurement_point, unique_fields=("vehicle_model_id", "code"), order_by=("code",)),
        _resource("master.factory-vehicle-models", "工厂-车型关系", "主数据关系", FactoryVehicleModel, FactoryVehicleModelCreate, None, bind_factory_vehicle_model, unique_fields=("factory_id", "vehicle_model_id"), order_by=("created_at",)),
        _resource("master.vehicle-model-colors", "车型-颜色关系", "主数据关系", VehicleModelColor, VehicleModelColorCreate, None, bind_vehicle_model_color, unique_fields=("vehicle_model_id", "color_id"), order_by=("created_at",)),
        _resource("master.measurement-group-points", "测量编组-点位关系", "主数据关系", MeasurementGroupPoint, MeasurementGroupPointBind, None, bind_measurement_group_point, unique_fields=("measurement_group_id", "measurement_point_id"), order_by=("sequence_no",)),
        _resource("process.parameter-definitions", "参数定义", "工艺", ParameterDefinition, ParameterDefinitionCreate, None, create_parameter_definition, unique_fields=("code",), order_by=("code",)),
        _resource("process.parameter-constraint-sources", "参数约束来源", "工艺", ParameterConstraintSource, ParameterConstraintSourceCreate, ParameterConstraintSourceUpdate, create_parameter_constraint_source, update_parameter_constraint_source, unique_fields=("constraint_code",), order_by=("constraint_code",)),
        _resource("process.spray-programs", "喷涂程序", "工艺", SprayProgram, SprayProgramCreate, SprayProgramUpdate, create_spray_program, update_spray_program, unique_fields=("factory_id", "program_code"), order_by=("program_code",)),
        _resource("process.program-versions", "喷涂程序版本", "工艺", SprayProgramVersion, SprayProgramVersionCreate, SprayProgramVersionUpdate, create_program_version, update_program_version, unique_fields=("spray_program_id", "version"), create_arg_fields=("spray_program_id",), extra_fields=(BulkField("spray_program_id", "喷涂程序 ID", "string", True, "填 spray_program.id", "创建版本所需父级程序 ID"),), export_row=_program_version_row, order_by=("created_at",)),
        _resource("process.program-vehicle-models", "程序适用车型", "工艺关系", ProgramVehicleModel, ProgramVehicleModelBulkCreate, None, _bind_program_vehicle_model, unique_fields=("program_version_id", "vehicle_model_id"), order_by=("created_at",)),
        _resource("process.program-colors", "程序适用颜色", "工艺关系", ProgramColor, ProgramColorBulkCreate, None, _bind_program_color, unique_fields=("program_version_id", "color_id"), order_by=("created_at",)),
        _resource("process.brushes", "刷子表/刷子号", "工艺", Brush, BrushCreate, BrushUpdate, create_brush, update_brush, unique_fields=("program_version_id", "brush_no"), create_arg_fields=("program_version_id",), extra_fields=(BulkField("program_version_id", "程序版本 ID", "string", True),), order_by=("brush_no",)),
        _resource("process.brush-parameters", "刷子参数", "工艺", BrushParameter, BrushParameterCreate, BrushParameterUpdate, create_brush_parameter, update_brush_parameter, unique_fields=("brush_id", "parameter_code"), create_arg_fields=("brush_id",), extra_fields=(BulkField("brush_id", "刷子 ID", "string", True),), order_by=("parameter_code",)),
        _resource("process.brush-contributions", "刷子点位贡献", "工艺", BrushPointContribution, BrushPointContributionUpsert, BrushPointContributionUpsert, upsert_brush_point_contribution, upsert_brush_point_contribution, create_arg_fields=("brush_id", "measurement_point_id"), update_arg_fields=("brush_id", "measurement_point_id"), update_uses_resource_id=False, extra_fields=(BulkField("brush_id", "刷子 ID", "string", True), BulkField("measurement_point_id", "测量点 ID", "string", True)), match_existing=_brush_contribution_match, order_by=("created_at",)),
        _resource("process.material-batches", "材料批次", "生产", MaterialBatch, MaterialBatchCreate, MaterialBatchUpdate, create_material_batch, update_material_batch, unique_fields=("batch_no",), order_by=("batch_no",)),
        _resource("process.production-runs", "生产事件", "生产", ProductionRun, ProductionRunCreate, ProductionRunUpdate, create_production_run, update_production_run, unique_fields=("run_no",), order_by=("started_at",)),
        _resource("process.production-stage-runs", "生产工序实绩", "生产", ProductionStageRun, ProductionStageRunCreate, ProductionStageRunUpdate, create_production_stage_run, update_production_stage_run, unique_fields=("production_run_id", "process_stage"), create_arg_fields=("production_run_id",), extra_fields=(BulkField("production_run_id", "生产事件 ID", "string", True),), order_by=("created_at",)),
        _resource("process.actual-parameters", "实际参数", "生产", ActualParameter, ActualParameterCreate, ActualParameterUpdate, create_actual_parameter, update_actual_parameter, create_arg_fields=("production_stage_run_id",), extra_fields=(BulkField("production_stage_run_id", "生产工序实绩 ID", "string", True),), order_by=("sampled_at",)),
        _resource("quality.measurements", "质量测量", "质量", QualityMeasurement, QualityMeasurementCreate, QualityMeasurementUpdate, create_quality_measurement, update_quality_measurement, unique_fields=("data_no",), order_by=("measured_at",), export_row=_quality_measurement_row),
        _resource("quality.standards", "质量标准", "质量", QualityStandard, QualityStandardCreate, QualityStandardUpdate, create_quality_standard, update_quality_standard, unique_fields=("standard_no", "version", "quality_type", "metric_code"), order_by=("standard_no",)),
        _resource("quality.body-map-layouts", "检测点二维位置", "质量点位", MeasurementPointLayout, BodyMapLayoutUpsert, BodyMapLayoutUpsert, upsert_body_map_layout, upsert_body_map_layout, unique_fields=("measurement_point_id", "body_view"), create_arg_fields=("measurement_point_id",), update_arg_fields=("measurement_point_id",), update_uses_resource_id=False, extra_fields=(BulkField("measurement_point_id", "检测点 ID", "string", True, "填 measurement_point.id", "检测点的逻辑引用"),), order_by=("body_view", "created_at")),
        _resource("quality.body-map-3d-layouts", "检测点三维位置", "质量点位", MeasurementPoint3DLayout, BodyMap3DLayoutUpsert, BodyMap3DLayoutUpsert, upsert_body_map_3d_layout, upsert_body_map_3d_layout, unique_fields=("measurement_point_id",), create_arg_fields=("measurement_point_id",), update_arg_fields=("measurement_point_id",), update_uses_resource_id=False, extra_fields=(BulkField("measurement_point_id", "检测点 ID", "string", True, "填 measurement_point.id", "检测点的逻辑引用"),), export_row=_body_map_3d_export_row, order_by=("created_at",)),
        _resource("measurement-governance.instruments", "测量仪器", "仪器治理", MeasurementInstrument, MeasurementInstrumentCreate, MeasurementInstrumentUpdate, create_instrument, update_instrument, unique_fields=("code",), order_by=("code",)),
        _resource("measurement-governance.methods", "测量方法", "仪器治理", MeasurementMethod, MeasurementMethodCreate, MeasurementMethodUpdate, create_measurement_method, update_measurement_method, unique_fields=("code", "version"), order_by=("code", "version")),
        _resource("measurement-governance.references", "参考件", "仪器治理", MeasurementReferenceStandard, MeasurementReferenceStandardCreate, MeasurementReferenceStandardUpdate, create_reference, update_reference, unique_fields=("code",), order_by=("code",)),
        _resource("measurement-governance.import-profiles", "仪器导入模板", "仪器治理", MeasurementImportProfile, MeasurementImportProfileCreate, MeasurementImportProfileUpdate, create_import_profile, update_import_profile, unique_fields=("code", "version"), order_by=("code", "version")),
        _resource("measurement-governance.calibrations", "校准/检查记录", "仪器治理", MeasurementCalibrationRecord, MeasurementCalibrationCreate, MeasurementCalibrationUpdate, create_calibration, update_calibration, unique_fields=("calibration_no",), order_by=("calibrated_at",)),
        _resource("engineering.process-routes", "3C3B 工艺路线", "工程闭环", ProcessRoute, ProcessRouteCreate, ProcessRouteUpdate, create_process_route, update_process_route, unique_fields=("factory_id", "route_code", "version"), order_by=("route_code", "version")),
        _resource("engineering.process-route-steps", "3C3B 路线步骤", "工程闭环", ProcessRouteStep, ProcessRouteStepCreate, ProcessRouteStepUpdate, create_process_route_step, update_process_route_step, unique_fields=("process_route_id", "step_code"), order_by=("process_route_id", "sequence_no")),
        _resource("engineering.process-route-applicabilities", "路线适用车型颜色", "工程闭环", ProcessRouteApplicability, ProcessRouteApplicabilityCreate, ProcessRouteApplicabilityUpdate, create_process_route_applicability, update_process_route_applicability, unique_fields=("process_route_id", "vehicle_model_id", "color_id"), order_by=("created_at",)),
        _resource("engineering.file-import-profiles", "设备/材料文件导入 Profile", "工程闭环", FileImportProfile, FileImportProfileCreate, FileImportProfileUpdate, create_file_import_profile, update_file_import_profile, unique_fields=("code", "version"), order_by=("domain_type", "code", "version")),
        _resource("engineering.file-import-jobs", "设备/材料文件导入任务", "工程闭环", FileImportJob, FileImportJobCreate, FileImportJobUpdate, reject_direct_file_import_job_create, update_file_import_job, unique_fields=("import_no",), order_by=("submitted_at",), importable=False),
        _resource("engineering.measurement-probes", "测量探头", "工程闭环", MeasurementProbe, MeasurementProbeCreate, MeasurementProbeUpdate, create_measurement_probe, update_measurement_probe, unique_fields=("instrument_id", "code"), order_by=("instrument_id", "code")),
        _resource("engineering.measurement-msa-studies", "测量 MSA/GRR", "工程闭环", MeasurementMsaStudy, MeasurementMsaStudyCreate, MeasurementMsaStudyUpdate, create_measurement_msa_study, update_measurement_msa_study, unique_fields=("study_no",), order_by=("study_at",)),
        _resource("engineering.issue-tasks", "质量问题/调试工单", "工程闭环", QualityIssueTask, QualityIssueTaskCreate, QualityIssueTaskUpdate, create_issue_task, update_issue_task, unique_fields=("task_no",), order_by=("created_at",)),
        _resource("engineering.issue-evidence", "问题工单证据", "工程闭环", QualityIssueEvidence, QualityIssueEvidenceCreate, None, create_issue_task_evidence, unique_fields=("task_id", "source_type", "source_id", "evidence_type"), create_arg_fields=("task_id",), extra_fields=(BulkField("task_id", "问题工单 ID", "string", True),), order_by=("created_at",)),
        _resource("engineering.issue-comments", "问题工单协作记录", "工程闭环", QualityIssueComment, QualityIssueCommentCreate, None, create_issue_task_comment, create_arg_fields=("task_id",), extra_fields=(BulkField("task_id", "问题工单 ID", "string", True),), order_by=("created_at",)),
        _resource("engineering.knowledge-entries", "诊断知识库", "工程闭环", EngineeringKnowledgeEntry, EngineeringKnowledgeEntryCreate, EngineeringKnowledgeEntryUpdate, create_knowledge_entry, update_knowledge_entry, unique_fields=("entry_code", "version"), order_by=("category", "entry_code", "version")),
        _resource("engineering.supplier-submissions", "供应商材料提交", "工程闭环", SupplierMaterialSubmission, SupplierMaterialSubmissionCreate, SupplierMaterialSubmissionUpdate, create_supplier_submission, update_supplier_submission, unique_fields=("submission_no",), order_by=("submitted_at",)),
        _resource("engineering.supplier-issues", "供应商材料问题", "工程闭环", SupplierMaterialIssue, SupplierMaterialIssueCreate, SupplierMaterialIssueUpdate, create_supplier_issue, update_supplier_issue, unique_fields=("issue_no",), order_by=("created_at",)),
        _resource("engineering.contribution-validations", "点位贡献验证", "工程闭环", ContributionValidationStudy, ContributionValidationStudyCreate, ContributionValidationStudyUpdate, create_contribution_validation, update_contribution_validation, unique_fields=("contribution_version_id", "study_no"), order_by=("created_at",)),
        _resource("engineering.trajectory-geometries", "轨迹段几何治理", "工程闭环", TrajectorySegmentGeometry, TrajectorySegmentGeometryCreate, TrajectorySegmentGeometryUpdate, create_trajectory_geometry, update_trajectory_geometry, unique_fields=("path_segment_id", "geometry_version"), order_by=("created_at",)),
        _resource("engineering.model-explanations", "模型解释/SHAP", "工程闭环", ModelExplanation, ModelExplanationCreate, None, create_model_explanation, unique_fields=("model_version_id", "prediction_result_id", "explanation_type", "target_metric"), order_by=("generated_at",)),
        _resource("material-governance.definitions", "材料特性定义", "材料治理", MaterialCharacteristicDefinition, MaterialCharacteristicDefinitionCreate, MaterialCharacteristicDefinitionUpdate, create_material_definition, update_material_definition, unique_fields=("code",), order_by=("code",)),
        _resource("material-governance.methods", "材料检测方法", "材料治理", MaterialTestMethod, MaterialTestMethodCreate, MaterialTestMethodUpdate, create_material_method, update_material_method, unique_fields=("code", "version"), order_by=("code", "version")),
        _resource("material-governance.specifications", "材料规格", "材料治理", MaterialSpecification, MaterialSpecificationCreate, MaterialSpecificationUpdate, create_material_specification, update_material_specification, unique_fields=("material_code", "characteristic_definition_id", "method_id", "version"), order_by=("material_code", "created_at")),
        _resource("material-governance.applicabilities", "材料适用关系", "材料治理", MaterialCharacteristicApplicability, MaterialCharacteristicApplicabilityCreate, MaterialCharacteristicApplicabilityUpdate, create_material_applicability, update_material_applicability, unique_fields=("characteristic_definition_id", "material_type", "process_stage", "target_family"), order_by=("process_stage",)),
        _resource("material-governance.results", "材料批次检测结果", "材料治理", MaterialBatchTestResult, MaterialBatchTestResultCreate, MaterialBatchTestResultUpdate, create_material_result, update_material_result, unique_fields=("result_no",), order_by=("tested_at",)),
        _resource("robot-governance.robots", "Dürr 机器人", "Dürr 治理", DurrRobot, DurrRobotCreate, DurrRobotUpdate, create_robot, update_robot, unique_fields=("factory_id", "code"), order_by=("code",)),
        _resource("robot-governance.controllers", "Dürr 应用控制器", "Dürr 治理", DurrApplicationController, DurrControllerCreate, DurrControllerUpdate, create_controller, update_controller, unique_fields=("factory_id", "code"), order_by=("code",)),
        _resource("robot-governance.atomizers", "Dürr 静电旋杯", "Dürr 治理", DurrRotaryAtomizer, DurrAtomizerCreate, DurrAtomizerUpdate, create_atomizer, update_atomizer, unique_fields=("factory_id", "code"), order_by=("code",)),
        _resource("robot-governance.device-configurations", "程序设备配置", "Dürr 治理", ProgramDeviceConfiguration, ProgramDeviceConfigurationCreate, ProgramDeviceConfigurationUpdate, create_device_configuration, update_device_configuration, unique_fields=("program_version_id", "configuration_version"), order_by=("created_at",)),
        _resource("robot-governance.trajectory-programs", "轨迹程序", "Dürr 治理", TrajectoryProgram, TrajectoryProgramCreate, TrajectoryProgramUpdate, create_trajectory_program, update_trajectory_program, unique_fields=("program_version_id", "trajectory_code", "version"), order_by=("created_at",)),
        _resource("robot-governance.path-segments", "轨迹路径段", "Dürr 治理", TrajectoryPathSegment, TrajectoryPathSegmentCreate, TrajectoryPathSegmentUpdate, create_path_segment, update_path_segment, unique_fields=("trajectory_program_id", "segment_no"), order_by=("trajectory_program_id", "segment_no")),
        _resource("robot-governance.contribution-versions", "点位贡献版本", "Dürr 治理", PointContributionVersion, PointContributionVersionCreate, PointContributionVersionUpdate, create_contribution_version, update_contribution_version, unique_fields=("program_version_id", "target_family", "version"), order_by=("created_at",)),
        _resource("robot-governance.contribution-entries", "点位贡献条目", "Dürr 治理", PointContributionEntry, PointContributionEntryCreate, PointContributionEntryUpdate, create_contribution_entry, update_contribution_entry, match_existing=_contribution_entry_match, order_by=("contribution_version_id", "measurement_point_id")),
        _resource("robot-governance.device-executions", "设备轨迹执行", "Dürr 治理", ProductionDeviceExecution, ProductionDeviceExecutionCreate, ProductionDeviceExecutionUpdate, create_device_execution, update_device_execution, unique_fields=("production_stage_run_id",), order_by=("created_at",)),
        _resource("integrations.endpoints", "集成端点", "集成", IntegrationEndpoint, IntegrationEndpointCreate, IntegrationEndpointUpdate, create_endpoint, update_endpoint, unique_fields=("code",), order_by=("code",)),
        _resource("integrations.events", "集成事件", "集成", IntegrationEvent, IntegrationEventCreate, None, create_event, unique_fields=("endpoint_id", "source_event_id"), order_by=("created_at",)),
    ]
}
