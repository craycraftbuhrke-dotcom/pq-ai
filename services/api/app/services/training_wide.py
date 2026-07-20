from __future__ import annotations

import csv
import hashlib
from datetime import UTC, date, datetime
from io import BytesIO, StringIO
from typing import Any
from uuid import uuid4
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile

from fastapi import HTTPException, Response
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.domain.scope_policy import (
    CURRENT_FEATURE_SET_VERSION,
    ScopeViolation,
    require_scope_safe_model,
    target_family_for_metric,
)
from app.models.domain import (
    MaterialCharacteristicDefinition,
    ParameterDefinition,
    PointFeatureSnapshot,
    TrainingDataUpload,
    TrainingWideSample,
)

IDENTITY_COLUMNS = ("样本编号", "独立分组", "样本时间", "目标值")

STAGES = (
    ("midcoat", "中涂外喷"),
    ("basecoat_1", "色漆一站"),
    ("basecoat_2", "色漆二站"),
    ("clearcoat_1", "清漆一站"),
    ("clearcoat_2", "清漆二站"),
)

PARAMETERS = (
    ("spray_flow", "喷涂流量"),
    ("bell_speed", "旋杯转速"),
    ("outer_air", "外成型空气流量"),
    ("inner_air", "内成型空气流量"),
    ("voltage", "静电高压"),
    ("gun_distance", "喷枪距离"),
    ("gun_spacing", "喷枪间距"),
    ("spray_speed", "喷涂速度"),
)

MATERIAL_PARAMETERS = (
    ("viscosity", "材料粘度"),
    ("solid_ratio", "材料固含比"),
)


def _default_feature_labels() -> dict[str, str]:
    labels: dict[str, str] = {}
    for stage_code, stage_label in STAGES:
        for parameter_code, parameter_label in PARAMETERS:
            labels[f"{stage_code}.{parameter_code}"] = f"{stage_label}-{parameter_label}"
        for parameter_code, parameter_label in MATERIAL_PARAMETERS:
            labels[f"{stage_code}.material.{parameter_code}"] = (
                f"{stage_label}-{parameter_label}"
            )
    return labels


def _parameter_name_map(db: Session) -> dict[str, str]:
    names = {
        item.code: item.name
        for item in db.scalars(select(ParameterDefinition).order_by(ParameterDefinition.code))
    }
    names.update(
        {
            item.code: item.name
            for item in db.scalars(
                select(MaterialCharacteristicDefinition).order_by(
                    MaterialCharacteristicDefinition.code
                )
            )
        }
    )
    return names


def _humanize_feature(feature_name: str, names: dict[str, str]) -> str:
    defaults = _default_feature_labels()
    if feature_name in defaults:
        return defaults[feature_name]
    parts = feature_name.split(".")
    stage_labels = dict(STAGES)
    if len(parts) >= 2 and parts[0] in stage_labels:
        parameter_code = parts[-1]
        parameter_name = names.get(parameter_code)
        if parameter_name:
            return f"{stage_labels[parts[0]]}-{parameter_name}"
    raise HTTPException(
        status_code=422,
        detail=f"特征 {feature_name} 尚未配置中文名称，不能生成面向用户的宽表",
    )


def feature_label_map(db: Session, target_metric: str) -> dict[str, str]:
    try:
        family = target_family_for_metric(target_metric)
    except ScopeViolation as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    feature_names = set(_default_feature_labels())
    snapshots = db.scalars(
        select(PointFeatureSnapshot).where(
            PointFeatureSnapshot.feature_set_version == CURRENT_FEATURE_SET_VERSION,
            PointFeatureSnapshot.target_family == family,
        )
    )
    for snapshot in snapshots:
        feature_names.update(str(key) for key in (snapshot.feature_values or {}))
    names = _parameter_name_map(db)
    result: dict[str, str] = {}
    used_labels: dict[str, str] = {}
    for feature_name in sorted(feature_names):
        label = _humanize_feature(feature_name, names)
        previous = used_labels.get(label)
        if previous and previous != feature_name:
            raise HTTPException(
                status_code=422,
                detail=f"训练宽表中文列名 {label} 对应多个参数，请先完善参数名称治理",
            )
        result[feature_name] = label
        used_labels[label] = feature_name
    return result


def _spreadsheet_safe_value(value: Any) -> Any:
    if isinstance(value, str) and value.lstrip(" \t\r\n").startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _csv_bytes(columns: list[str], rows: list[dict[str, Any]]) -> bytes:
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=columns)
    writer.writeheader()
    for row in rows:
        writer.writerow(
            {
                column: _spreadsheet_safe_value(row.get(column, ""))
                for column in columns
            }
        )
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def _xlsx_bytes(columns: list[str], rows: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "训练数据"
    sheet.append(columns)
    for row in rows:
        sheet.append(
            [_spreadsheet_safe_value(row.get(column, "")) for column in columns]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(columns)).column_letter}1"
    for index, column in enumerate(columns, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = min(
            max(len(column) + 4, 14), 28
        )

    instructions = workbook.create_sheet("填写说明")
    instructions.append(["项目", "说明"])
    instructions.append(["样本编号", "每行唯一，例如企业内部试验编号或车身编号"])
    instructions.append(["独立分组", "同一车身、批次或同一次试验必须填写相同分组，防止数据泄漏"])
    instructions.append(["样本时间", "填写实际发生时间，系统按时间划分训练和验证数据"])
    instructions.append(["目标值", "填写本次训练要预测的质量指标实测值"])
    instructions.append(["工艺参数", "只填写真实记录值；空白表示该样本没有该项，不要填写猜测值"])
    instructions.append(["数据范围", "仅支持中涂、色漆、清漆及膜厚、色差、橘皮相关数据"])
    instructions.column_dimensions["A"].width = 18
    instructions.column_dimensions["B"].width = 80
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def training_template_response(
    db: Session, target_metric: str, file_format: str
) -> Response:
    columns = [*IDENTITY_COLUMNS, *feature_label_map(db, target_metric).values()]
    if file_format == "xlsx":
        content = _xlsx_bytes(columns, [])
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        suffix = "xlsx"
    else:
        content = _csv_bytes(columns, [])
        media_type = "text/csv; charset=utf-8"
        suffix = "csv"
    return Response(
        content=content,
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="training-wide-template.{suffix}"'
        },
    )


def _parse_rows(content: bytes, filename: str) -> tuple[list[str], list[dict[str, Any]]]:
    if not content:
        raise HTTPException(status_code=422, detail="训练数据文件为空")
    if filename.lower().endswith(".xlsx"):
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except (BadZipFile, InvalidFileException, KeyError, OSError, ParseError, ValueError) as exc:
            raise HTTPException(
                status_code=422,
                detail="Excel 文件无法解析，请确认文件格式正确且未损坏",
            ) from exc
        sheet = workbook["训练数据"] if "训练数据" in workbook.sheetnames else workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration:
            raise HTTPException(status_code=422, detail="Excel 缺少表头") from None
        headers = [str(value).strip() if value is not None else "" for value in raw_headers]
        parsed = []
        for values in rows:
            if not values or all(value is None or str(value).strip() == "" for value in values):
                continue
            parsed.append(
                {
                    header: values[index] if index < len(values) else None
                    for index, header in enumerate(headers)
                }
            )
        return headers, parsed
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise HTTPException(status_code=422, detail="CSV 必须使用 UTF-8 编码") from exc
    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(status_code=422, detail="CSV 缺少表头")
    headers = [str(value).strip() for value in reader.fieldnames]
    return headers, [dict(row) for row in reader]


def _required_text(row: dict[str, Any], column: str, row_number: int) -> str:
    raw = row.get(column, "")
    if isinstance(raw, float) and raw.is_integer():
        raw = int(raw)
    value = str(raw or "").strip()
    if not value:
        raise ValueError(f"第 {row_number} 行「{column}」不能为空")
    return value


def _number(value: Any, label: str, row_number: int) -> float:
    if value is None or str(value).strip() == "":
        raise ValueError(f"第 {row_number} 行「{label}」不能为空")
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"第 {row_number} 行「{label}」必须是数字") from exc


def _datetime(value: Any, row_number: int) -> datetime:
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, datetime.min.time())
    else:
        text = str(value or "").strip()
        if not text:
            raise ValueError(f"第 {row_number} 行「样本时间」不能为空")
        try:
            parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError as exc:
            raise ValueError(
                f"第 {row_number} 行「样本时间」格式不正确，请使用 YYYY-MM-DD HH:MM:SS"
            ) from exc
    return parsed.replace(tzinfo=UTC) if parsed.tzinfo is None else parsed.astimezone(UTC)


def validate_training_file(
    db: Session,
    content: bytes,
    filename: str,
    target_metric: str,
    feature_set_version: str,
) -> dict[str, Any]:
    try:
        require_scope_safe_model(target_metric, feature_set_version, [])
    except ScopeViolation as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    headers, rows = _parse_rows(content, filename)
    duplicate_headers = sorted(
        {header for header in headers if header and headers.count(header) > 1}
    )
    if duplicate_headers:
        raise HTTPException(
            status_code=422,
            detail=f"存在重复列：{', '.join(duplicate_headers)}；请每个参数只保留一列",
        )
    missing = [column for column in IDENTITY_COLUMNS if column not in headers]
    if missing:
        raise HTTPException(status_code=422, detail=f"缺少必填列：{', '.join(missing)}")
    label_to_feature = {
        label: feature for feature, label in feature_label_map(db, target_metric).items()
    }
    feature_headers = [header for header in headers if header not in IDENTITY_COLUMNS]
    unknown = [header for header in feature_headers if header not in label_to_feature]
    if unknown:
        raise HTTPException(
            status_code=422,
            detail=f"存在不认识的列：{', '.join(unknown[:10])}；请使用系统提供的最新模板",
        )
    if not feature_headers:
        raise HTTPException(status_code=422, detail="训练宽表至少需要一个工艺或材料参数列")
    mapped_features = [label_to_feature[header] for header in feature_headers]
    duplicate_features = sorted(
        {feature for feature in mapped_features if mapped_features.count(feature) > 1}
    )
    if duplicate_features:
        raise HTTPException(status_code=422, detail="存在重复的工艺或材料参数列")
    if not rows:
        raise HTTPException(status_code=422, detail="训练宽表没有数据行")

    samples: list[dict[str, Any]] = []
    errors: list[str] = []
    sample_numbers: set[str] = set()
    missing_value_count = 0
    for row_number, row in enumerate(rows, start=2):
        try:
            sample_no = _required_text(row, "样本编号", row_number)
            if sample_no in sample_numbers:
                raise ValueError(f"第 {row_number} 行样本编号 {sample_no} 重复")
            sample_numbers.add(sample_no)
            feature_values: dict[str, float] = {}
            for header in feature_headers:
                raw = row.get(header)
                if raw is None or str(raw).strip() == "":
                    missing_value_count += 1
                    continue
                feature_values[label_to_feature[header]] = _number(raw, header, row_number)
            if not feature_values:
                raise ValueError(f"第 {row_number} 行没有填写任何工艺或材料参数")
            samples.append(
                {
                    "sample_no": sample_no,
                    "group_value": _required_text(row, "独立分组", row_number),
                    "occurred_at": _datetime(row.get("样本时间"), row_number),
                    "target_value": _number(row.get("目标值"), "目标值", row_number),
                    "feature_values": feature_values,
                    "row_number": row_number,
                }
            )
        except ValueError as exc:
            errors.append(str(exc))
    if errors:
        raise HTTPException(
            status_code=422,
            detail={"message": "训练宽表校验未通过", "errors": errors[:100]},
        )
    feature_names = sorted({key for sample in samples for key in sample["feature_values"]})
    try:
        require_scope_safe_model(target_metric, feature_set_version, feature_names)
    except ScopeViolation as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return {
        "samples": samples,
        "feature_names": feature_names,
        "report": {
            "passed": True,
            "sample_count": len(samples),
            "group_count": len({sample["group_value"] for sample in samples}),
            "feature_count": len(feature_names),
            "blank_feature_cell_count": missing_value_count,
            "errors": [],
        },
    }


def import_training_file(
    db: Session,
    content: bytes,
    filename: str,
    upload_name: str,
    target_metric: str,
    feature_set_version: str,
    uploaded_by: str,
) -> TrainingDataUpload:
    validated = validate_training_file(
        db, content, filename, target_metric, feature_set_version
    )
    now = datetime.now(UTC)
    digest = hashlib.sha256(content).hexdigest()
    upload = TrainingDataUpload(
        upload_no=f"MAN-{now:%Y%m%d%H%M%S}-{uuid4().hex[:8].upper()}",
        name=upload_name.strip(),
        target_metric=target_metric,
        feature_set_version=feature_set_version,
        source_type="MANUAL_UPLOAD",
        file_name=filename,
        file_hash=digest,
        status="VALIDATED",
        sample_count=len(validated["samples"]),
        feature_names=validated["feature_names"],
        validation_report=validated["report"],
        uploaded_by=uploaded_by,
        uploaded_at=now,
    )
    db.add(upload)
    db.flush()
    for sample in validated["samples"]:
        db.add(
            TrainingWideSample(
                upload_id=upload.id,
                sample_no=sample["sample_no"],
                group_value=sample["group_value"],
                occurred_at=sample["occurred_at"],
                target_value=sample["target_value"],
                feature_values=sample["feature_values"],
                lineage={
                    "source_type": "MANUAL_UPLOAD",
                    "file_name": filename,
                    "file_hash": digest,
                    "row_number": sample["row_number"],
                },
                is_valid=True,
            )
        )
    db.commit()
    db.refresh(upload)
    return upload


def training_upload_export_response(
    db: Session, upload: TrainingDataUpload, file_format: str
) -> Response:
    labels = feature_label_map(db, upload.target_metric)
    missing_labels = [name for name in upload.feature_names if name not in labels]
    if missing_labels:
        raise HTTPException(
            status_code=422,
            detail=(
                f"该训练记录包含 {len(missing_labels)} 个已停用的参数列，"
                "请由管理员恢复参数中文名称后再导出"
            ),
        )
    columns = [*IDENTITY_COLUMNS, *[labels[name] for name in upload.feature_names]]
    samples = list(
        db.scalars(
            select(TrainingWideSample)
            .where(TrainingWideSample.upload_id == upload.id)
            .order_by(TrainingWideSample.occurred_at, TrainingWideSample.sample_no)
        )
    )
    rows = []
    for sample in samples:
        row: dict[str, Any] = {
            "样本编号": sample.sample_no,
            "独立分组": sample.group_value,
            "样本时间": sample.occurred_at.isoformat(),
            "目标值": sample.target_value,
        }
        row.update(
            {
                labels[name]: sample.feature_values.get(name, "")
                for name in upload.feature_names
            }
        )
        rows.append(row)
    if file_format == "xlsx":
        content = _xlsx_bytes(columns, rows)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        suffix = "xlsx"
    else:
        content = _csv_bytes(columns, rows)
        media_type = "text/csv; charset=utf-8"
        suffix = "csv"
    return Response(
        content=content,
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="training-wide-{upload.upload_no}.{suffix}"'
        },
    )
