"""Tests for the recipe-wide template/export/import endpoints."""

from io import BytesIO, StringIO

import asyncio

from openpyxl import load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

from app.api.routes.recipe_wide import (
    recipe_wide_export,
    recipe_wide_import,
    recipe_wide_template,
)
from app.models.domain import (
    Brush,
    BrushParameter,
    BrushPointContribution,
    Factory,
    MeasurementPoint,
    Part,
    ProcessStage,
    SprayProgram,
    SprayProgramVersion,
    VehicleModel,
)
from tests.schema_guard import create_transient_test_schema


def _build_session() -> Session:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    create_transient_test_schema(engine)
    return Session(engine)


def _seed(db: Session) -> tuple[str, str, str]:
    factory = Factory(code="F1", name="Factory 1")
    db.add(factory)
    vm = VehicleModel(code="V1", name="Vehicle 1")
    db.add(vm)
    part = Part(code="DOOR", name="Door")
    db.add(part)
    db.flush()
    program = SprayProgram(
        program_code="PROG-1",
        name="Program 1",
        factory_id=factory.id,
        process_stage=ProcessStage.CLEARCOAT_2.value,
        station_code="STN-1",
        station_name="Station 1",
    )
    db.add(program)
    db.flush()
    version = SprayProgramVersion(
        spray_program_id=program.id,
        version="v1",
    )
    db.add(version)
    point = MeasurementPoint(
        code="P1",
        name="Point 1",
        vehicle_model_id=vm.id,
        part_id=part.id,
        point_type="QUALITY",
    )
    db.add(point)
    db.flush()
    db.commit()
    return program.id, version.id, point.id


def _csv_bytes(columns: list[str], rows: list[dict]) -> bytes:
    import csv

    buf = StringIO()
    writer = csv.DictWriter(buf, fieldnames=columns)
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def _run_import(db, program_id, version_id, content):
    class _Req:
        async def body(self) -> bytes:
            return content

    return asyncio.run(recipe_wide_import(
        _Req(),
        spray_program_id=program_id,
        program_version_id=version_id,
        brush_table_no="T1",
        filename="recipe-wide-import.csv",
        db=db,
    ))


def test_template_empty_returns_prefilled_row() -> None:
    db = _build_session()
    try:
        program_id, version_id, _ = _seed(db)
        response = recipe_wide_template(
            spray_program_id=program_id,
            program_version_id=version_id,
            brush_table_no=None,
            file_format="csv",
            db=db,
        )
        assert response.status_code == 200
        text = response.body.decode("utf-8-sig")
        lines = text.strip().splitlines()
        header = lines[0].split(",")
        assert "喷涂程序" in header
        assert "喷涂流量" in header
        assert "喷涂流量_下限" in header
        assert "权重" in header
        # one prefilled blank row
        assert len(lines) == 2
        assert "PROG-1" in lines[1]
        assert "v1" in lines[1]
    finally:
        db.close()


def test_export_with_data_emits_rows() -> None:
    db = _build_session()
    try:
        program_id, version_id, point_id = _seed(db)
        brush = Brush(
            program_version_id=version_id,
            brush_no="B1",
            brush_table_no="T1",
            spray_position="roof",
        )
        db.add(brush)
        db.flush()
        db.add(
            BrushParameter(
                brush_id=brush.id,
                parameter_code="clearcoat_2_spray_flow",
                parameter_name="清漆二站喷涂流量",
                configured_value=320.0,
                unit="ml/min",
                soft_min=280.0,
                soft_max=360.0,
            )
        )
        db.add(
            BrushPointContribution(
                brush_id=brush.id,
                measurement_point_id=point_id,
                overlap_ratio=0.5,
                contribution_weight=0.6,
            )
        )
        db.commit()
        response = recipe_wide_export(
            spray_program_id=program_id,
            program_version_id=version_id,
            brush_table_no="T1",
            file_format="xlsx",
            db=db,
        )
        assert response.status_code == 200
        workbook = load_workbook(BytesIO(response.body), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        header = list(rows[0])
        assert "喷涂流量" in header
        assert "权重" in header
        data_rows = [r for r in rows[1:] if any(c not in (None, "") for c in r)]
        assert len(data_rows) == 1
    finally:
        db.close()


def test_import_creates_chain() -> None:
    db = _build_session()
    try:
        program_id, version_id, _point_id = _seed(db)
        columns = [
            "喷涂程序", "版本", "刷子表", "刷子号", "喷涂点位", "测量点",
            "喷涂流量", "喷涂流量_下限", "喷涂流量_上限",
            "外成型空气流量", "外成型空气流量_下限", "外成型空气流量_上限",
            "内成型空气流量", "内成型空气流量_下限", "内成型空气流量_上限",
            "旋杯转速", "旋杯转速_下限", "旋杯转速_上限",
            "静电高压", "静电高压_下限", "静电高压_上限",
            "权重",
        ]
        rows = [{
            "喷涂程序": "PROG-1", "版本": "v1", "刷子表": "T1", "刷子号": "B1",
            "喷涂点位": "roof", "测量点": "P1",
            "喷涂流量": "320", "喷涂流量_下限": "280", "喷涂流量_上限": "360",
            "外成型空气流量": "", "外成型空气流量_下限": "", "外成型空气流量_上限": "",
            "内成型空气流量": "", "内成型空气流量_下限": "", "内成型空气流量_上限": "",
            "旋杯转速": "", "旋杯转速_下限": "", "旋杯转速_上限": "",
            "静电高压": "", "静电高压_下限": "", "静电高压_上限": "",
            "权重": "0.6",
        }]
        content = _csv_bytes(columns, rows)
        result = _run_import(db, program_id, version_id, content)
        assert result.failed == 0, result.errors
        assert result.created >= 1
        brush = db.scalar(select(Brush).where(Brush.brush_no == "B1"))
        assert brush is not None
        assert brush.brush_table_no == "T1"
        param = db.scalar(
            select(BrushParameter).where(BrushParameter.brush_id == brush.id)
        )
        assert param is not None
        assert param.configured_value == 320.0
        assert param.soft_min == 280.0
        assert param.soft_max == 360.0
        contrib = db.scalar(
            select(BrushPointContribution).where(
                BrushPointContribution.brush_id == brush.id
            )
        )
        assert contrib is not None
        assert contrib.contribution_weight == 0.6
    finally:
        db.close()


def test_import_updates_existing() -> None:
    db = _build_session()
    try:
        program_id, version_id, point_id = _seed(db)
        brush = Brush(program_version_id=version_id, brush_no="B1", brush_table_no="T1")
        db.add(brush)
        db.flush()
        db.add(
            BrushParameter(
                brush_id=brush.id,
                parameter_code="clearcoat_2_spray_flow",
                parameter_name="清漆二站喷涂流量",
                configured_value=300.0,
                unit="ml/min",
            )
        )
        db.commit()
        columns = [
            "喷涂程序", "版本", "刷子表", "刷子号", "喷涂点位", "测量点",
            "喷涂流量", "喷涂流量_下限", "喷涂流量_上限",
            "外成型空气流量", "外成型空气流量_下限", "外成型空气流量_上限",
            "内成型空气流量", "内成型空气流量_下限", "内成型空气流量_上限",
            "旋杯转速", "旋杯转速_下限", "旋杯转速_上限",
            "静电高压", "静电高压_下限", "静电高压_上限",
            "权重",
        ]
        rows = [{
            "喷涂程序": "PROG-1", "版本": "v1", "刷子表": "T1", "刷子号": "B1",
            "喷涂点位": "", "测量点": "P1",
            "喷涂流量": "350", "喷涂流量_下限": "", "喷涂流量_上限": "",
            "外成型空气流量": "", "外成型空气流量_下限": "", "外成型空气流量_上限": "",
            "内成型空气流量": "", "内成型空气流量_下限": "", "内成型空气流量_上限": "",
            "旋杯转速": "", "旋杯转速_下限": "", "旋杯转速_上限": "",
            "静电高压": "", "静电高压_下限": "", "静电高压_上限": "",
            "权重": "0.7",
        }]
        content = _csv_bytes(columns, rows)
        result = _run_import(db, program_id, version_id, content)
        assert result.failed == 0, result.errors
        param = db.scalar(
            select(BrushParameter).where(BrushParameter.brush_id == brush.id)
        )
        assert param.configured_value == 350.0
        contrib = db.scalar(
            select(BrushPointContribution).where(
                BrushPointContribution.brush_id == brush.id
            )
        )
        assert contrib is not None
        assert contrib.contribution_weight == 0.7
    finally:
        db.close()


def test_import_bad_point_records_error() -> None:
    db = _build_session()
    try:
        program_id, version_id, _ = _seed(db)
        columns = [
            "喷涂程序", "版本", "刷子表", "刷子号", "喷涂点位", "测量点",
            "喷涂流量", "喷涂流量_下限", "喷涂流量_上限",
            "外成型空气流量", "外成型空气流量_下限", "外成型空气流量_上限",
            "内成型空气流量", "内成型空气流量_下限", "内成型空气流量_上限",
            "旋杯转速", "旋杯转速_下限", "旋杯转速_上限",
            "静电高压", "静电高压_下限", "静电高压_上限",
            "权重",
        ]
        rows = [{
            "喷涂程序": "PROG-1", "版本": "v1", "刷子表": "T1", "刷子号": "B1",
            "喷涂点位": "", "测量点": "NOPE",
            "喷涂流量": "320", "喷涂流量_下限": "", "喷涂流量_上限": "",
            "外成型空气流量": "", "外成型空气流量_下限": "", "外成型空气流量_上限": "",
            "内成型空气流量": "", "内成型空气流量_下限": "", "内成型空气流量_上限": "",
            "旋杯转速": "", "旋杯转速_下限": "", "旋杯转速_上限": "",
            "静电高压": "", "静电高压_下限": "", "静电高压_上限": "",
            "权重": "0.6",
        }]
        content = _csv_bytes(columns, rows)
        result = _run_import(db, program_id, version_id, content)
        assert result.failed == 1
        assert "测量点 NOPE" in result.errors[0].message
    finally:
        db.close()
