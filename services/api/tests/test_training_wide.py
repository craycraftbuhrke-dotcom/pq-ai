from datetime import UTC, datetime, timedelta
from io import BytesIO

import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.domain.scope_policy import CURRENT_FEATURE_SET_VERSION
from app.models import domain  # noqa: F401
from app.models.domain import (
    Color,
    DatasetSplitMember,
    Factory,
    FactoryVehicleModel,
    MeasurementPoint,
    Part,
    PointFeatureSnapshot,
    ProductionRun,
    QualityMeasurement,
    QualityMetricValue,
    TrainingWideSample,
    VehicleModel,
    VehicleModelColor,
)
from app.schemas.modeling import DatasetBuildRequest, ModelTrainingRequest
from app.services.modeling import build_dataset_snapshot, train_model
from app.services.training_wide import (
    import_training_file,
    training_template_response,
    training_upload_export_response,
    validate_training_file,
)
from tests.schema_guard import create_transient_test_schema


def _engine():
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    create_transient_test_schema(engine)
    return engine


def _seed_context(
    db: Session,
    *,
    factory_code: str = "F1",
    vehicle_code: str = "V1",
    color_code: str = "C1",
) -> tuple[Factory, VehicleModel, Color]:
    factory = Factory(code=factory_code, name=f"工厂{factory_code}")
    vehicle = VehicleModel(code=vehicle_code, name=f"车型{vehicle_code}")
    color = Color(code=color_code, name=f"颜色{color_code}", color_type="BASECOAT")
    db.add_all([factory, vehicle, color])
    db.flush()
    db.add_all(
        [
            FactoryVehicleModel(
                factory_id=factory.id, vehicle_model_id=vehicle.id, is_active=True
            ),
            VehicleModelColor(
                vehicle_model_id=vehicle.id, color_id=color.id, is_active=True
            ),
        ]
    )
    db.flush()
    return factory, vehicle, color


def _manual_csv(
    now: datetime,
    *,
    factory_code: str = "F1",
    vehicle_code: str = "V1",
    color_code: str = "C1",
    row_count: int = 6,
) -> bytes:
    rows = [
        "样本编号,独立分组,工厂代码,车型代码,颜色代码,样本时间,目标值,"
        "清漆二站-喷涂流量,清漆二站-外成型空气流量"
    ]
    for index in range(row_count):
        rows.append(
            f"S-{index},G-{index},{factory_code},{vehicle_code},{color_code},"
            f"{(now + timedelta(hours=index)).isoformat()},"
            f"{10 + index},{100 + index},{200 + index}"
        )
    return ("\ufeff" + "\n".join(rows)).encode("utf-8")


def test_training_wide_rejects_broken_excel_and_duplicate_columns() -> None:
    with Session(_engine(), expire_on_commit=False) as db:
        with pytest.raises(HTTPException) as broken:
            validate_training_file(
                db,
                b"not-an-excel-workbook",
                "broken.xlsx",
                "doi",
                CURRENT_FEATURE_SET_VERSION,
            )
        assert broken.value.status_code == 422
        assert "无法解析" in str(broken.value.detail)

        duplicate_csv = (
            "样本编号,样本编号,独立分组,工厂代码,车型代码,颜色代码,样本时间,目标值,清漆二站-喷涂流量\n"
            "S-1,S-2,G-1,F1,V1,C1,2026-07-19 08:00:00,10,100\n"
        ).encode("utf-8")
        with pytest.raises(HTTPException) as duplicate:
            validate_training_file(
                db,
                duplicate_csv,
                "duplicate.csv",
                "doi",
                CURRENT_FEATURE_SET_VERSION,
            )
        assert duplicate.value.status_code == 422
        assert "重复列" in str(duplicate.value.detail)


def test_training_wide_normalizes_numeric_sample_number_and_escapes_exports() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "训练数据"
    sheet.append(
        [
            "样本编号",
            "独立分组",
            "工厂代码",
            "车型代码",
            "颜色代码",
            "样本时间",
            "目标值",
            "清漆二站-喷涂流量",
        ]
    )
    sheet.append(
        [100.0, "+UNSAFE-GROUP", "F1", "V1", "C1", datetime(2026, 7, 19, 8, 0), 10, 100]
    )
    source = BytesIO()
    workbook.save(source)

    with Session(_engine(), expire_on_commit=False) as db:
        _seed_context(db)
        upload = import_training_file(
            db,
            source.getvalue(),
            "manual.xlsx",
            "人工训练数据",
            "doi",
            CURRENT_FEATURE_SET_VERSION,
            "测试上传人",
        )
        sample = db.scalar(
            select(TrainingWideSample).where(TrainingWideSample.upload_id == upload.id)
        )
        assert sample.sample_no == "100"
        assert sample.factory_id is not None
        assert sample.vehicle_model_id is not None
        assert sample.color_id is not None

        csv_response = training_upload_export_response(db, upload, "csv")
        assert "'+UNSAFE-GROUP" in csv_response.body.decode("utf-8-sig")
        assert "F1" in csv_response.body.decode("utf-8-sig")

        xlsx_response = training_upload_export_response(db, upload, "xlsx")
        exported = load_workbook(BytesIO(xlsx_response.body), read_only=True, data_only=False)
        assert exported["训练数据"]["B2"].value == "'+UNSAFE-GROUP"


def test_manual_training_wide_can_build_dataset_without_production_rows() -> None:
    now = datetime.now(UTC) - timedelta(days=1)
    with Session(_engine(), expire_on_commit=False) as db:
        factory, vehicle, color = _seed_context(db)
        template = training_template_response(db, "doi", "xlsx")
        assert "spreadsheetml" in template.media_type
        assert len(template.body) > 1000
        upload = import_training_file(
            db,
            _manual_csv(now),
            "manual-training.csv",
            "人工受控试验训练记录",
            "doi",
            CURRENT_FEATURE_SET_VERSION,
            "测试上传人",
        )
        assert upload.sample_count == 6
        assert upload.validation_report["passed"] is True
        assert upload.validation_report["context_counts"] == {
            "factory": 1,
            "vehicle_model": 1,
            "color": 1,
        }
        samples = list(
            db.scalars(
                select(TrainingWideSample).where(
                    TrainingWideSample.upload_id == upload.id
                )
            )
        )
        assert all(sample.factory_id == factory.id for sample in samples)
        assert all(sample.vehicle_model_id == vehicle.id for sample in samples)
        assert all(sample.color_id == color.id for sample in samples)
        dataset = build_dataset_snapshot(
            db,
            DatasetBuildRequest(
                dataset_code="MANUAL-ONLY",
                version="1.0",
                target_metric="doi",
                include_all_production=False,
                manual_upload_ids=[upload.id],
                min_train_groups=3,
                min_validation_groups=2,
            ),
        )
        assert dataset.sample_count == 6
        assert dataset.lineage["source_policy"] == "PRODUCTION_AND_MANUAL_EQUAL_WEIGHT"
        assert dataset.lineage["source_counts"] == {
            "PRODUCTION": 0,
            "MANUAL_UPLOAD": 6,
        }
        members = list(
            db.scalars(
                select(DatasetSplitMember).where(
                    DatasetSplitMember.dataset_snapshot_id == dataset.id
                )
            )
        )
        assert all(member.source_type == "MANUAL_UPLOAD" for member in members)
        assert all(member.production_run_id is None for member in members)
        assert all(member.manual_sample_id for member in members)

        trained = train_model(
            db,
            ModelTrainingRequest(
                model_code="MANUAL-DOI",
                version="1.0",
                target_metric="doi",
                dataset_snapshot_id=dataset.id,
                min_samples=3,
            ),
        )
        validation = trained.evaluation_metrics["multi_axis_validation"]["axes"]
        assert validation["TIME_HOLDOUT"]["validation_sample_count"] == 2
        assert validation["PRODUCTION_GROUP_LOO"]["status"] == "EVALUATED"
        assert validation["PRODUCTION_GROUP_LOO"]["validation_sample_count"] == 6
        # 人工样本已带生产同语义上下文；单工厂时轴有数据但多样性不足
        assert validation["FACTORY"]["excluded_sample_count"] == 0
        assert validation["FACTORY"]["status"] == "INSUFFICIENT_AXIS_DIVERSITY"
        assert validation["COLOR"]["excluded_sample_count"] == 0
        assert trained.model_payload["uncertainty_source"] == "TEMPORAL_VALIDATION_RMSE"


def test_production_and_manual_rows_enter_same_matrix_without_source_priority() -> None:
    now = datetime.now(UTC) - timedelta(days=2)
    with Session(_engine(), expire_on_commit=False) as db:
        factory, model, color = _seed_context(
            db, factory_code="F", vehicle_code="M", color_code="C"
        )
        part = Part(code="P", name="零件")
        db.add(part)
        db.flush()
        point = MeasurementPoint(
            code="MP",
            name="点位",
            vehicle_model_id=model.id,
            part_id=part.id,
            quality_types=["ORANGE_PEEL"],
        )
        db.add(point)
        db.flush()
        production_snapshot_ids = []
        for index in range(4):
            run = ProductionRun(
                run_no=f"R-{index}",
                body_no=f"B-{index}",
                factory_id=factory.id,
                vehicle_model_id=model.id,
                color_id=color.id,
                started_at=now + timedelta(hours=index),
            )
            db.add(run)
            db.flush()
            snapshot = PointFeatureSnapshot(
                production_run_id=run.id,
                measurement_point_id=point.id,
                feature_set_version=CURRENT_FEATURE_SET_VERSION,
                target_family="ORANGE_PEEL",
                feature_values={
                    "clearcoat_2.spray_flow": 90 + index,
                    "clearcoat_2.outer_air": 190 + index,
                },
                completeness_score=1,
                generated_at=run.started_at,
            )
            measurement = QualityMeasurement(
                data_no=f"Q-{index}",
                production_run_id=run.id,
                measurement_point_id=point.id,
                quality_type="ORANGE_PEEL",
                measured_at=run.started_at,
                reliability_status="VERIFIED",
            )
            db.add_all([snapshot, measurement])
            db.flush()
            db.add(
                QualityMetricValue(
                    measurement_id=measurement.id,
                    metric_code="doi",
                    metric_name="DOI",
                    raw_value=20 + index,
                )
            )
            production_snapshot_ids.append(snapshot.id)
        db.commit()
        upload = import_training_file(
            db,
            _manual_csv(
                now + timedelta(days=1),
                factory_code="F",
                vehicle_code="M",
                color_code="C",
            ),
            "mixed-training.csv",
            "混合训练记录",
            "doi",
            CURRENT_FEATURE_SET_VERSION,
            "测试上传人",
        )
        dataset = build_dataset_snapshot(
            db,
            DatasetBuildRequest(
                dataset_code="MIXED",
                version="1.0",
                target_metric="doi",
                include_all_production=False,
                production_snapshot_ids=production_snapshot_ids,
                manual_upload_ids=[upload.id],
                min_train_groups=3,
                min_validation_groups=2,
            ),
        )
        assert dataset.sample_count == 10
        assert dataset.lineage["source_counts"] == {
            "PRODUCTION": 4,
            "MANUAL_UPLOAD": 6,
        }
        assert dataset.feature_names == [
            "clearcoat_2.outer_air",
            "clearcoat_2.spray_flow",
        ]
        manual_samples = list(
            db.scalars(
                select(TrainingWideSample).where(TrainingWideSample.upload_id == upload.id)
            )
        )
        assert len(manual_samples) == 6
        assert all(sample.factory_id == factory.id for sample in manual_samples)
